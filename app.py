import streamlit as st
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import pandas as pd
import time, re, json, hashlib, datetime
from collections import deque, Counter
from io import BytesIO
from difflib import SequenceMatcher

# textstat pour lisibilité (pip install textstat)
try:
    import textstat
    TEXTSTAT_OK = True
except ImportError:
    TEXTSTAT_OK = False

st.set_page_config(page_title="Scopo - SEO Audit", page_icon="🔍", layout="wide",
    menu_items={"Get help":None,"Report a bug":None,"About":None})

st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');

/* ── RESET STREAMLIT CHROME ── */
#MainMenu{display:none !important}
footer{display:none !important}
header{display:none !important}
[data-testid="stToolbar"]{display:none !important}
[data-testid="stHeader"]{display:none !important}
[data-testid="stDecoration"]{display:none !important}
.stDeployButton{display:none !important}
button[kind="header"]{display:none !important}

/* ── BASE ── */
.stApp{background:#F8F9FA;font-family:'Inter',sans-serif}
.block-container{padding-top:1rem;padding-bottom:2rem}

/* ── HEADER SCOPO ── */
.app-header{background:#FFFFFF;padding:1rem 2rem;border-radius:0;margin-bottom:1.8rem;border-bottom:1px solid #E5E7EB}
.app-header h1{color:#111827 !important;margin:0;font-size:1.5rem;font-weight:800;letter-spacing:-.02em}
.app-header p{color:#6B7280;margin:3px 0 0;font-size:.85rem}

/* ── KPI CARDS ── */
.kpi-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:1.2rem}
.kpi-card{background:#FFFFFF;border-radius:12px;padding:1rem 1.1rem;box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB}
.kpi-val{font-size:2rem;font-weight:800;color:#111827;line-height:1.1}
.kpi-lbl{font-size:.75rem;color:#6B7280;margin-top:3px;font-weight:500;text-transform:uppercase;letter-spacing:.04em}
.kpi-card.orange .kpi-val::before{content:"● ";color:#2563EB;font-size:1rem}
.kpi-card.red .kpi-val::before{content:"● ";color:#EF4444;font-size:1rem}
.kpi-card.green .kpi-val::before{content:"● ";color:#10B981;font-size:1rem}
.kpi-card.blue .kpi-val::before{content:"● ";color:#2563EB;font-size:1rem}
.kpi-card.teal .kpi-val::before{content:"● ";color:#2563EB;font-size:1rem}
.kpi-card.purple .kpi-val::before{content:"● ";color:#7C3AED;font-size:1rem}

/* ── SCORE GLOBAL ── */
.score-wrap{background:#FFFFFF;border-radius:12px;padding:1.5rem;text-align:center;border:1px solid #E5E7EB}
.score-big{font-size:5rem;font-weight:900;line-height:1}
.score-sub{font-size:.9rem;color:#6B7280;margin-top:2px}
.score-lbl{font-size:1rem;font-weight:700;margin-top:6px}
.score-bar{height:6px;border-radius:8px;background:#F3F4F6;margin-top:12px;overflow:hidden}
.score-bar-fill{height:100%;border-radius:8px}

/* ── SUGGESTIONS ── */
.sug{border-radius:8px;padding:.9rem 1.1rem;margin-bottom:.7rem;border-left:3px solid}
.sug-title{font-weight:600;color:#111827;font-size:.9rem}
.sug-action{color:#6B7280;font-size:.83rem;margin-top:4px}
.sug-chip{display:inline-block;font-size:.7rem;font-weight:600;padding:2px 8px;border-radius:20px;margin-top:6px}
.stProgress>div>div{background:#2563EB !important}

/* ── SIDEBAR ── */
section[data-testid="stSidebar"]{background:#FFFFFF; min-width:280px !important; width:300px !important}
section[data-testid="stSidebar"] > div{padding:1rem 1rem 2rem 1rem !important}

section[data-testid="stSidebar"] h3{
    font-size:.75rem !important; font-weight:600 !important;
    color:#111827 !important; letter-spacing:.08em; text-transform:uppercase;
    margin-top:.5rem !important; margin-bottom:.3rem !important;
    white-space:nowrap !important; overflow:visible !important;
}

section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] .stSlider label,
section[data-testid="stSidebar"] .stCheckbox label,
section[data-testid="stSidebar"] .stToggle label,
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"],
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p{
    font-size:.82rem !important; color:#374151 !important;
    font-weight:500 !important; white-space:nowrap !important;
    overflow:visible !important; text-overflow:clip !important; min-width:0 !important;
}

section[data-testid="stSidebar"] .stCheckbox{background:transparent !important}
section[data-testid="stSidebar"] .stCheckbox span{color:#374151 !important; font-size:.82rem !important}
section[data-testid="stSidebar"] .stCheckbox input[type="checkbox"]{accent-color:#2563EB}
section[data-testid="stSidebar"] .stToggle span{color:#374151 !important; font-size:.82rem !important}

section[data-testid="stSidebar"] input[type="text"],
section[data-testid="stSidebar"] input[type="password"]{
    font-size:.82rem !important; background:#fff !important;
    color:#111827 !important; border:1px solid #E5E7EB !important; border-radius:8px !important;
}

section[data-testid="stSidebar"] .stSlider [data-testid="stTickBar"]{display:none}
section[data-testid="stSidebar"] .stSlider .stMarkdown{font-size:.75rem !important; color:#6B7280 !important}

section[data-testid="stSidebar"] .stButton button{
    background:#2563EB !important; color:white !important;
    font-weight:700 !important; border:none !important;
    border-radius:8px !important; font-size:.9rem !important;
}
section[data-testid="stSidebar"] .stButton button:hover{background:#1D4ED8 !important}

section[data-testid="stSidebar"] hr{margin:.4rem 0 !important; border-color:#F3F4F6 !important}

/* ── ONGLETS ── */
.stTabs [data-baseweb="tab-list"]{
    gap:4px !important; background:#F3F4F6 !important; border-radius:8px !important;
    padding:4px !important; flex-wrap:nowrap !important; overflow-x:auto !important;
}
.stTabs [data-baseweb="tab"]{
    background:transparent !important; border-radius:7px !important;
    padding:6px 14px !important; font-size:.8rem !important; font-weight:500 !important;
    color:#6B7280 !important; white-space:nowrap !important; border:none !important; min-width:fit-content !important;
}
.stTabs [aria-selected="true"]{background:white !important; color:#111827 !important; box-shadow:0 1px 4px rgba(0,0,0,.08) !important}
.stTabs [data-baseweb="tab"]:hover{color:#111827 !important; background:rgba(255,255,255,.6) !important}
.stTabs [data-baseweb="tab-highlight"]{background:#2563EB !important; height:2px !important; border-radius:2px !important}

/* ── LISIBILITÉ ── */
[data-testid="stMetricValue"]{color:#111827 !important;font-weight:700 !important}
[data-testid="stMetricLabel"]{color:#6B7280 !important}
[data-testid="stMetricDelta"]{color:#374151 !important}
div[data-testid="stAlert"] p{color:#111827 !important}
div[data-testid="stAlert"][kind="warning"] p{color:#92400e !important}
div[data-testid="stAlert"][kind="info"] p{color:#1e40af !important}
div[data-testid="stAlert"][kind="success"] p{color:#065f46 !important}
div[data-testid="stAlert"][kind="error"] p{color:#991b1b !important}
.stAlert p{color:#111827 !important}
[data-baseweb="notification"] span{color:#111827 !important}

.stTabs .stMarkdown p{color:#111827 !important}
[data-testid="stVerticalBlock"] .stMarkdown p{color:#111827 !important}
.stTabs .stMarkdown h1,.stTabs .stMarkdown h2,.stTabs .stMarkdown h3,
.stTabs .stMarkdown h4,.stTabs .stMarkdown h5,.stTabs .stMarkdown h6{color:#111827 !important}

[data-testid="stCaptionContainer"] p{color:#6B7280 !important}
div[data-testid="stMarkdownContainer"] blockquote{border-left:3px solid #2563EB !important;background:#EFF6FF !important;padding:.5rem 1rem !important;border-radius:4px !important}
div[data-testid="stMarkdownContainer"] blockquote p{color:#1E40AF !important;font-weight:500 !important}
div[data-testid="stMarkdownContainer"] blockquote *{color:#1E40AF !important}

/* ── BOUTONS DOWNLOAD ── */
.stDownloadButton button{
    background:#FFFFFF !important; color:#111827 !important;
    font-weight:600 !important; border:1px solid #E5E7EB !important;
    border-radius:8px !important; font-size:.92rem !important; padding:.7rem 1rem !important;
}
.stDownloadButton button:hover{border-color:#2563EB !important; color:#2563EB !important}
</style>""", unsafe_allow_html=True)

HEADERS = {"User-Agent":"Mozilla/5.0 (compatible; ScopoBot/1.0; +https://scopo.app)","Accept-Language":"fr-FR,fr;q=0.9"}

FILE_TYPES = {
    "pdf":   [".pdf"],
    "image": [".jpg",".jpeg",".png",".gif",".webp",".svg",".ico"],
    "doc":   [".doc",".docx",".xls",".xlsx",".ppt",".pptx",".odt"],
    "media": [".mp4",".mp3",".avi",".mov",".wav"],
    "zip":   [".zip",".rar",".gz",".tar"],
    "code":  [".js",".css",".json",".xml"],
}

def get_domain(url):
    p=urlparse(url); return f"{p.scheme}://{p.netloc}"

def is_internal(url,base):
    return urlparse(url).netloc==urlparse(base).netloc

def normalize_url(url):
    p=urlparse(url); return f"{p.scheme}://{p.netloc}{p.path}".rstrip("/")

def get_file_type(url):
    path=urlparse(url).path.lower()
    for ftype,exts in FILE_TYPES.items():
        if any(path.endswith(e) for e in exts): return ftype
    return "html"

def fetch(url,timeout=10):
    try:
        t0=time.time()
        r=requests.get(url,headers=HEADERS,timeout=timeout,allow_redirects=True)
        return r,int((time.time()-t0)*1000)
    except: return None,0

def fetch_head(url,timeout=8):
    try:
        r=requests.head(url,headers=HEADERS,timeout=timeout,allow_redirects=True)
        size=r.headers.get("Content-Length","")
        return r.status_code, int(size) if size.isdigit() else 0
    except: return 0,0

def should_exclude(url,excl_ext,excl_pat):
    p=urlparse(url); path=p.path.lower()
    if any(path.endswith(e) for e in excl_ext): return True
    if any(pat.lower() in url.lower() for pat in excl_pat): return True
    return False

def text_hash(text):
    return hashlib.md5(re.sub(r'\s+',' ',text.strip().lower())[:500].encode()).hexdigest()

def analyze_page(url,resp,elapsed):
    r={"url":url,"status":resp.status_code if resp else 0,"redirected_to":"",
       "file_type":"html","file_size_kb":0,
       "title":"","title_len":0,"h1":"","h1_count":0,
       "meta_description":"","meta_desc_len":0,"meta_noindex":False,
       "canonical":"","canonical_self":True,
       "images_total":0,"images_no_alt":0,
       "images_non_webp":0,"images_heavy":0,
       "links_internal":0,"links_external":0,
       "outbound_domains":[],"outbound_links_detail":[],
       "word_count":0,"load_time_ms":elapsed,
       "internal_links_list":[],"body_text_hash":"","body_text_sample":"",
       "og_title":"","og_description":"","og_image":"",
       "has_schema":False,"hreflang_count":0,
       "h2_count":0,"h3_count":0,"depth":0,
       "cta_empty":0,"links_broken_anchor":0,
       "flesch_score":None,
       "render_blocking_js":0,"render_blocking_css":0,
       "accessibility_score":None,
       "issues":[],"seo_score":0}

    if not resp: r["issues"].append("Erreur réseau"); return r
    if resp.url!=url: r["redirected_to"]=resp.url
    if resp.status_code!=200:
        r["issues"].append(f"HTTP {resp.status_code}"); return r
    try: soup=BeautifulSoup(resp.text,"html.parser")
    except: r["issues"].append("Erreur parsing"); return r

    # Title
    t=soup.find("title")
    if t:
        tv=t.get_text(strip=True); r["title"]=tv; r["title_len"]=len(tv)
        if not tv: r["issues"].append("Title vide")
        elif len(tv)>60: r["issues"].append(f"Title long ({len(tv)})")
        elif len(tv)<30: r["issues"].append(f"Title court ({len(tv)})")
    else: r["issues"].append("Title manquant")

    # Hn
    h1s=soup.find_all("h1"); r["h1_count"]=len(h1s)
    if h1s: r["h1"]=h1s[0].get_text(strip=True)[:120]
    if not h1s: r["issues"].append("H1 manquant")
    elif len(h1s)>1: r["issues"].append(f"{len(h1s)} H1")
    r["h2_count"]=len(soup.find_all("h2"))
    r["h3_count"]=len(soup.find_all("h3"))

    # Méta
    md=soup.find("meta",attrs={"name":re.compile("^description$",re.I)})
    if md:
        dv=md.get("content",""); r["meta_description"]=dv[:160]; r["meta_desc_len"]=len(dv)
        if not dv: r["issues"].append("Méta vide")
        elif len(dv)>160: r["issues"].append(f"Méta longue ({len(dv)})")
        elif len(dv)<70: r["issues"].append(f"Méta courte ({len(dv)})")
    else: r["issues"].append("Méta manquante")

    # Robots
    rm=soup.find("meta",attrs={"name":re.compile("^robots$",re.I)})
    if rm and "noindex" in rm.get("content","").lower():
        r["meta_noindex"]=True; r["issues"].append("⚠ noindex")

    # Canonical
    can=soup.find("link",rel="canonical")
    if can:
        ch=can.get("href",""); r["canonical"]=ch
        if ch and normalize_url(ch)!=normalize_url(url):
            r["canonical_self"]=False; r["issues"].append("Canonical ailleurs")

    # Open Graph
    for prop,key in [("og:title","og_title"),("og:description","og_description"),("og:image","og_image")]:
        m=soup.find("meta",property=prop) or soup.find("meta",attrs={"name":prop})
        if m: r[key]=m.get("content","")[:200]

    # Schema.org
    r["has_schema"]=bool(soup.find("script",type="application/ld+json"))

    # Hreflang
    r["hreflang_count"]=len(soup.find_all("link",rel="alternate",hreflang=True))

    # Images
    imgs=soup.find_all("img"); r["images_total"]=len(imgs)
    na=[i for i in imgs if not i.get("alt","").strip()]
    r["images_no_alt"]=len(na)
    if na: r["issues"].append(f"{len(na)} img sans ALT")

    # Images non WebP / lourdes (src extension)
    MODERN_FMTS = (".webp", ".svg", ".avif")
    non_webp = []
    for i in imgs:
        src = i.get("src", "").strip()
        if not src or src.startswith("data:"): continue  # skip inline base64
        src_clean = src.split("?")[0].lower()  # ignore query params
        if not any(src_clean.endswith(e) for e in MODERN_FMTS):
            non_webp.append(i)
    r["images_non_webp"] = len(non_webp)
    if non_webp: r["issues"].append(f"{len(non_webp)} img non WebP")

    # CTA / boutons vides (boutons et liens sans texte)
    buttons=soup.find_all(["button","a"])
    empty_cta=[b for b in buttons if not b.get_text(strip=True) and not b.get("aria-label","").strip()]
    r["cta_empty"]=len(empty_cta)
    if len(empty_cta)>0: r["issues"].append(f"{len(empty_cta)} CTA vide(s)")

    # Liens
    base=get_domain(url)
    int_links=[]; out_links=[]; out_domains=set()
    for a in soup.find_all("a",href=True):
        href=a["href"].strip()
        if href.startswith(("#","mailto:","tel:")): continue
        full=urljoin(url,href)
        if urlparse(full).scheme not in ("http","https"): continue
        anchor=a.get_text(strip=True)[:80]
        if is_internal(full,base):
            int_links.append(normalize_url(full))
        else:
            dom=urlparse(full).netloc
            out_links.append({"url":full,"domain":dom,"anchor":anchor})
            out_domains.add(dom)
    r["links_internal"]=len(set(int_links))
    r["links_external"]=len(out_links)
    r["internal_links_list"]=list(set(int_links))
    r["outbound_domains"]=list(out_domains)
    r["outbound_links_detail"]=out_links[:50]  # max 50 par page

    # Profondeur
    r["depth"]=url.replace(base,"").count("/")

    # Contenu
    for s in soup(["script","style","nav","footer","header"]): s.decompose()
    text=soup.get_text(separator=" ")
    words=[w for w in text.split() if len(w)>2]
    r["word_count"]=len(words)
    body=" ".join(words)
    r["body_text_hash"]=text_hash(body)
    r["body_text_sample"]=body[:300]
    if len(words)<100: r["issues"].append(f"Court ({len(words)} mots)")

    # Score de lisibilité Flesch (FR)
    if TEXTSTAT_OK and len(words)>30:
        try:
            textstat.set_lang("fr")
            fs=textstat.flesch_reading_ease(body)
            r["flesch_score"]=round(fs,1)
            if fs<30: r["issues"].append(f"Lisibilité difficile (Flesch {fs:.0f})")
        except: pass

    r["seo_score"]=score_page(r)
    return r

def score_page(r):
    s=100
    if r.get("status",200)!=200: return 0
    if not r.get("title"): s-=20
    elif r.get("title_len",0)>60: s-=8
    elif r.get("title_len",0)<30: s-=5
    if not r.get("h1"): s-=15
    elif r.get("h1_count",1)>1: s-=8
    if not r.get("meta_description"): s-=12
    elif r.get("meta_desc_len",0)>160: s-=5
    if r.get("meta_noindex"): s-=15
    if r.get("images_no_alt",0)>0: s-=min(10,r["images_no_alt"]*2)
    if r.get("word_count",200)<100: s-=8
    if not r.get("canonical_self",True): s-=5
    return max(0,s)

def site_score(df):
    if df.empty: return 0
    avg=df["seo_score"].mean()
    penalty=min(30,len(df[df["status"]!=200])*5)
    return max(0,int(avg-penalty))

def detect_duplicates(results):
    dups=[]; seen={}
    for r in results:
        h=r.get("body_text_hash",""); u=r["url"]
        if h and h in seen:
            dups.append({"url_1":seen[h],"url_2":u,"type":"Contenu identique"})
        elif h: seen[h]=u
    samples=[(r["url"],r.get("body_text_sample","")) for r in results if len(r.get("body_text_sample",""))>50]
    checked=set()
    for i,(u1,t1) in enumerate(samples[:40]):
        for u2,t2 in samples[i+1:min(i+8,len(samples))]:
            pair=tuple(sorted([u1,u2]))
            if pair in checked: continue
            checked.add(pair)
            sim=SequenceMatcher(None,t1[:300],t2[:300]).ratio()
            if sim>0.85 and u1!=u2:
                dups.append({"url_1":u1,"url_2":u2,"type":f"Similaire ({int(sim*100)}%)"})
    return dups

def get_pagespeed(url,api_key,strategy="mobile"):
    if not api_key: return {}
    try:
        r=requests.get("https://www.googleapis.com/pagespeedonline/v5/runPagespeed",
            params={"url":url,"key":api_key,"strategy":strategy,
                    "category":["performance","accessibility"]},timeout=30)
        if r.status_code==200:
            data=r.json()
            cats=data.get("lighthouseResult",{}).get("categories",{})
            audits=data.get("lighthouseResult",{}).get("audits",{})

            # Render-blocking resources
            rb=audits.get("render-blocking-resources",{})
            rb_items=rb.get("details",{}).get("items",[])
            rb_count=len(rb_items)
            rb_savings=rb.get("numericValue",0)

            # Score accessibilité
            acc_score=cats.get("accessibility",{}).get("score")
            acc_score=int(acc_score*100) if acc_score is not None else None

            # Top audits accessibilité échoués
            acc_issues=[]
            for audit_id,audit in audits.items():
                if audit.get("score") is not None and audit.get("score")<1:
                    if any(x in audit_id for x in ["color-contrast","image-alt","button-name",
                        "link-name","label","aria","heading","document"]):
                        acc_issues.append({"id":audit_id,
                            "titre":audit.get("title",""),
                            "description":audit.get("description","")[:120]})

            return {"score":int(cats.get("performance",{}).get("score",0)*100),
                    "lcp":audits.get("largest-contentful-paint",{}).get("displayValue","N/A"),
                    "cls":audits.get("cumulative-layout-shift",{}).get("displayValue","N/A"),
                    "tbt":audits.get("total-blocking-time",{}).get("displayValue","N/A"),
                    "accessibility_score":acc_score,
                    "accessibility_issues":acc_issues[:10],
                    "render_blocking_count":rb_count,
                    "render_blocking_savings_ms":int(rb_savings),
                    "render_blocking_items":[i.get("url","")[:80] for i in rb_items[:10]]}
    except Exception as e: return {"error":str(e)}
    return {}

def get_open_pagerank(domain, api_key):
    """Open PageRank - gratuit, 100 req/jour, clé sur openpagerank.com"""
    if not api_key: return {}
    clean=domain.replace("https://","").replace("http://","").rstrip("/")
    try:
        r=requests.get("https://openpagerank.com/api/v1.0/getPageRank",
            params={"domains[]":clean},
            headers={"API-OPR":api_key},timeout=10)
        if r.status_code==200:
            data=r.json()
            if data.get("response") and data["response"][0].get("status_code")==200:
                item=data["response"][0]
                return {"rank":item.get("rank","N/A"),
                        "page_rank_decimal":round(item.get("page_rank_decimal",0),2),
                        "domain":clean}
    except Exception as e: return {"error":str(e)}
    return {}


def analyze_security(base_url, results):
    """Audit sécurité : headers HTTP, cookies, HTTPS, mixed content, formulaires"""
    sec = {
        "https": False,
        "hsts": False, "hsts_value": "",
        "csp": False, "csp_value": "",
        "x_frame": False, "x_frame_value": "",
        "x_content_type": False,
        "referrer_policy": False, "referrer_value": "",
        "permissions_policy": False,
        "cookies": [],
        "cookies_secure": 0, "cookies_httponly": 0,
        "cookies_samesite": 0, "cookies_no_secure": 0,
        "cookies_no_httponly": 0, "cookies_no_samesite": 0,
        "mixed_content_pages": [],
        "insecure_forms": [],
        "score": 0,
        "issues": [],
        "recommendations": []
    }

    # ── 1. Vérification HTTPS + headers de sécurité
    try:
        r = requests.get(base_url, headers=HEADERS, timeout=10, allow_redirects=True)
        hdrs = {k.lower(): v for k, v in r.headers.items()}

        sec["https"] = r.url.startswith("https://")
        if not sec["https"]:
            sec["issues"].append("Site non sécurisé (HTTP)")
            sec["recommendations"].append({"p":"🔴 Critique","t":"HTTPS non activé","a":"Installer un certificat SSL/TLS et forcer la redirection HTTP→HTTPS.","impact":"Données visiteurs non chiffrées - pénalité Google"})

        # HSTS
        hsts_val = hdrs.get("strict-transport-security", "")
        sec["hsts"] = bool(hsts_val)
        sec["hsts_value"] = hsts_val[:120]
        if sec["https"] and not sec["hsts"]:
            sec["issues"].append("HSTS absent")
            sec["recommendations"].append({"p":"🟠 Important","t":"HSTS manquant","a":"Ajouter Strict-Transport-Security: max-age=31536000; includeSubDomains","impact":"Attaques de downgrade possibles"})

        # CSP
        csp_val = hdrs.get("content-security-policy", "")
        sec["csp"] = bool(csp_val)
        sec["csp_value"] = csp_val[:200]
        if not sec["csp"]:
            sec["issues"].append("CSP absent")
            sec["recommendations"].append({"p":"🟠 Important","t":"Content-Security-Policy manquant","a":"Définir une politique CSP restrictive pour limiter les sources de scripts.","impact":"Risque XSS - injection de scripts malveillants"})

        # X-Frame-Options
        xfo = hdrs.get("x-frame-options", "")
        sec["x_frame"] = bool(xfo)
        sec["x_frame_value"] = xfo
        if not sec["x_frame"] and not sec["csp"]:
            sec["issues"].append("X-Frame-Options absent")
            sec["recommendations"].append({"p":"🟠 Important","t":"X-Frame-Options manquant","a":"Ajouter X-Frame-Options: SAMEORIGIN pour bloquer le clickjacking.","impact":"Risque clickjacking"})

        # X-Content-Type-Options
        xcto = hdrs.get("x-content-type-options", "")
        sec["x_content_type"] = xcto.lower() == "nosniff"
        if not sec["x_content_type"]:
            sec["issues"].append("X-Content-Type-Options absent")
            sec["recommendations"].append({"p":"🟡 Normal","t":"X-Content-Type-Options manquant","a":"Ajouter X-Content-Type-Options: nosniff","impact":"MIME sniffing - exécution non voulue de fichiers"})

        # Referrer-Policy
        rp = hdrs.get("referrer-policy", "")
        sec["referrer_policy"] = bool(rp)
        sec["referrer_value"] = rp
        if not sec["referrer_policy"]:
            sec["issues"].append("Referrer-Policy absent")
            sec["recommendations"].append({"p":"🟡 Normal","t":"Referrer-Policy manquant","a":"Ajouter Referrer-Policy: strict-origin-when-cross-origin","impact":"Fuite d'URL interne vers des sites tiers"})

        # Permissions-Policy
        pp = hdrs.get("permissions-policy", hdrs.get("feature-policy", ""))
        sec["permissions_policy"] = bool(pp)
        if not sec["permissions_policy"]:
            sec["recommendations"].append({"p":"🟡 Normal","t":"Permissions-Policy absent","a":"Restreindre l'accès caméra/micro/géolocalisation via Permissions-Policy.","impact":"APIs sensibles accessibles par défaut"})

        # ── 2. Analyse des cookies
        session_r = requests.Session()
        session_r.get(base_url, headers=HEADERS, timeout=10, allow_redirects=True)
        all_cookies = []
        for c in session_r.cookies:
            cookie_info = {
                "name": c.name,
                "domain": c.domain or urlparse(base_url).netloc,
                "secure": c.secure,
                "httponly": c.has_nonstandard_attr("httponly") or c.has_nonstandard_attr("HttpOnly"),
                "samesite": c.get_nonstandard_attr("samesite") or c.get_nonstandard_attr("SameSite") or "",
                "expires": str(c.expires) if c.expires else "Session",
                "path": c.path,
            }
            all_cookies.append(cookie_info)

        sec["cookies"] = all_cookies
        sec["cookies_secure"] = sum(1 for c in all_cookies if c["secure"])
        sec["cookies_httponly"] = sum(1 for c in all_cookies if c["httponly"])
        sec["cookies_samesite"] = sum(1 for c in all_cookies if c["samesite"])
        sec["cookies_no_secure"] = sum(1 for c in all_cookies if not c["secure"])
        sec["cookies_no_httponly"] = sum(1 for c in all_cookies if not c["httponly"])
        sec["cookies_no_samesite"] = sum(1 for c in all_cookies if not c["samesite"])

        if sec["cookies_no_secure"] > 0 and sec["https"]:
            sec["issues"].append(f"{sec['cookies_no_secure']} cookie(s) sans flag Secure")
            sec["recommendations"].append({"p":"🔴 Critique","t":f"{sec['cookies_no_secure']} cookie(s) sans flag Secure","a":"Ajouter le flag Secure sur tous les cookies pour qu'ils ne transitent que via HTTPS.","impact":"Cookies interceptables en HTTP"})
        if sec["cookies_no_httponly"] > 0:
            sec["issues"].append(f"{sec['cookies_no_httponly']} cookie(s) sans flag HttpOnly")
            sec["recommendations"].append({"p":"🟠 Important","t":f"{sec['cookies_no_httponly']} cookie(s) sans HttpOnly","a":"Ajouter HttpOnly pour bloquer l'accès aux cookies via JavaScript.","impact":"Cookies volables via XSS"})
        if sec["cookies_no_samesite"] > 0:
            sec["issues"].append(f"{sec['cookies_no_samesite']} cookie(s) sans SameSite")
            sec["recommendations"].append({"p":"🟡 Normal","t":f"{sec['cookies_no_samesite']} cookie(s) sans attribut SameSite","a":"Ajouter SameSite=Lax ou SameSite=Strict sur les cookies.","impact":"Risque CSRF - requêtes croisées non voulues"})

    except Exception as e:
        sec["issues"].append(f"Erreur analyse headers : {e}")

    # ── 3. Contenu mixte (HTTP resources sur pages HTTPS) + formulaires non sécurisés
    for row in results:
        if row.get("file_type") != "html": continue
        page_url = row["url"]
        resp, _ = fetch(page_url)
        if not resp or resp.status_code != 200: continue
        try:
            soup = BeautifulSoup(resp.text, "html.parser")
            # Mixed content
            mixed = []
            for tag, attr in [("img","src"),("script","src"),("link","href"),("iframe","src")]:
                for el in soup.find_all(tag, **{attr: True}):
                    src = el[attr]
                    if src.startswith("http://"):
                        mixed.append(src[:80])
            if mixed:
                sec["mixed_content_pages"].append({"url": page_url, "ressources": mixed[:5], "count": len(mixed)})

            # Formulaires non sécurisés
            for form in soup.find_all("form"):
                action = form.get("action", "")
                has_password = bool(form.find("input", type="password"))
                if action.startswith("http://") or (has_password and not page_url.startswith("https://")):
                    sec["insecure_forms"].append({"page": page_url, "action": action[:80], "has_password": has_password})
        except: continue

    if sec["mixed_content_pages"]:
        sec["issues"].append(f"{len(sec['mixed_content_pages'])} page(s) avec contenu mixte HTTP/HTTPS")
        sec["recommendations"].append({"p":"🔴 Critique","t":f"{len(sec['mixed_content_pages'])} page(s) avec contenu mixte","a":"Remplacer toutes les ressources HTTP par leur version HTTPS.","impact":"Avertissement navigateur + pénalité SEO"})
    if sec["insecure_forms"]:
        sec["issues"].append(f"{len(sec['insecure_forms'])} formulaire(s) non sécurisé(s)")
        sec["recommendations"].append({"p":"🔴 Critique","t":f"{len(sec['insecure_forms'])} formulaire(s) non sécurisé(s)","a":"Forcer HTTPS sur toutes les actions de formulaires, surtout les champs mot de passe.","impact":"Données utilisateurs transmises en clair"})

    # ── Score sécurité /100
    score = 100
    if not sec["https"]: score -= 30
    if not sec["hsts"]: score -= 15
    if not sec["csp"]: score -= 15
    if not sec["x_frame"]: score -= 10
    if not sec["x_content_type"]: score -= 8
    if not sec["referrer_policy"]: score -= 7
    if not sec["permissions_policy"]: score -= 5
    if sec["cookies_no_secure"] > 0: score -= 10
    if sec["cookies_no_httponly"] > 0: score -= 8
    if sec["cookies_no_samesite"] > 0: score -= 5
    if sec["mixed_content_pages"]: score -= 15
    if sec["insecure_forms"]: score -= 20
    sec["score"] = max(0, score)
    return sec

def analyze_robots(base_url):
    """Analyse robots.txt"""
    result={"url":f"{base_url}/robots.txt","status":0,"found":False,
            "disallow_count":0,"allow_count":0,"sitemap_refs":[],
            "disallowed_paths":[],"user_agents":[],"raw":"","issues":[]}
    try:
        r=requests.get(f"{base_url}/robots.txt",headers=HEADERS,timeout=10)
        result["status"]=r.status_code
        if r.status_code==200:
            result["found"]=True
            result["raw"]=r.text[:3000]
            lines=[l.strip() for l in r.text.splitlines()]
            current_ua="*"
            for line in lines:
                if not line or line.startswith("#"): continue
                if line.lower().startswith("user-agent:"):
                    ua=line.split(":",1)[1].strip()
                    if ua not in result["user_agents"]: result["user_agents"].append(ua)
                    current_ua=ua
                elif line.lower().startswith("disallow:"):
                    path=line.split(":",1)[1].strip()
                    if path: result["disallow_count"]+=1; result["disallowed_paths"].append(path)
                elif line.lower().startswith("allow:"):
                    result["allow_count"]+=1
                elif line.lower().startswith("sitemap:"):
                    sm=line.split(":",1)[1].strip()
                    if sm: result["sitemap_refs"].append(sm)
            if result["disallow_count"]==0:
                result["issues"].append("Aucune règle Disallow - tout est indexable")
            if not result["sitemap_refs"]:
                result["issues"].append("Aucun lien vers sitemap.xml dans robots.txt")
        else:
            result["issues"].append(f"robots.txt introuvable (HTTP {r.status_code})")
    except Exception as e:
        result["issues"].append(f"Erreur : {e}")
    return result

def analyze_sitemap(base_url, robots_data=None):
    """Analyse sitemap.xml - suit les refs robots.txt si dispo"""
    urls_to_try=[]
    if robots_data and robots_data.get("sitemap_refs"):
        urls_to_try=robots_data["sitemap_refs"][:3]
    urls_to_try += [f"{base_url}/sitemap.xml", f"{base_url}/sitemap_index.xml"]

    result={"found":False,"url":"","status":0,"url_count":0,
            "urls_sample":[],"is_index":False,"sub_sitemaps":[],"issues":[]}
    for sm_url in urls_to_try:
        try:
            r=requests.get(sm_url,headers=HEADERS,timeout=10)
            result["url"]=sm_url; result["status"]=r.status_code
            if r.status_code==200:
                result["found"]=True
                try:
                    from xml.etree import ElementTree as ET
                    root=ET.fromstring(r.text)
                    ns={"sm":"http://www.sitemaps.org/schemas/sitemap/0.9"}
                    # Sitemap index ?
                    sitemaps=root.findall(".//sm:sitemap/sm:loc",ns)
                    if sitemaps:
                        result["is_index"]=True
                        result["sub_sitemaps"]=[s.text for s in sitemaps[:10]]
                        result["url_count"]=len(sitemaps)
                    else:
                        locs=root.findall(".//sm:url/sm:loc",ns)
                        result["url_count"]=len(locs)
                        result["urls_sample"]=[l.text for l in locs[:5]]
                    if result["url_count"]==0:
                        result["issues"].append("Sitemap vide ou format non reconnu")
                except:
                    result["issues"].append("Impossible de parser le XML du sitemap")
                break
        except: continue
    if not result["found"]:
        result["issues"].append("Aucun sitemap trouvé (sitemap.xml, sitemap_index.xml)")
    return result

def serp_preview_html(title, meta, url):
    """Génère un aperçu SERP fidèle à Google (60 car title, 160 méta)"""
    MAX_T=60; MAX_M=160
    t_cut=title[:MAX_T]+"…" if len(title)>MAX_T else title
    m_cut=meta[:MAX_M]+"…" if len(meta)>MAX_M else meta
    t_color="#1a0dab"; t_warn="#ef4444"
    t_col=t_warn if len(title)>MAX_T or len(title)<30 else t_color
    m_col="#ef4444" if len(meta)>MAX_M or (meta and len(meta)<70) else "#4d5156"
    url_short=url.replace("https://","").replace("http://","")[:60]
    title_bar_pct=min(100,int(len(title)/MAX_T*100))
    meta_bar_pct=min(100,int(len(meta)/MAX_M*100)) if meta else 0
    title_bar_col="#059669" if len(title)<=MAX_T and len(title)>=30 else "#ef4444"
    meta_bar_col="#059669" if meta and len(meta)<=MAX_M and len(meta)>=70 else "#f8ba07" if meta and len(meta)<70 else "#ef4444" if meta else "#e5e7eb"
    return f"""<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:1.1rem 1.3rem;
        font-family:Arial,sans-serif;max-width:600px;box-shadow:0 1px 4px rgba(0,0,0,.06)'>
      <div style='font-size:.78rem;color:#202124;margin-bottom:2px'>{url_short}</div>
      <div style='font-size:1.1rem;font-weight:400;color:{t_col};line-height:1.3;margin-bottom:3px'>{t_cut or "(Aucun title)"}</div>
      <div style='font-size:.85rem;color:{m_col};line-height:1.5'>{m_cut or "<i style='color:#9ca3af'>Aucune méta-description</i>"}</div>
      <div style='margin-top:.8rem;display:flex;gap:1.2rem;font-size:.75rem;color:#6b7280'>
        <span>Title : <b style='color:{title_bar_col}'>{len(title)}/{MAX_T} car.</b>
          <span style='display:inline-block;width:60px;height:5px;background:#e5e7eb;border-radius:3px;vertical-align:middle;margin-left:4px'>
          <span style='display:block;width:{title_bar_pct}%;height:100%;background:{title_bar_col};border-radius:3px'></span></span>
        </span>
        <span>Méta : <b style='color:{meta_bar_col}'>{len(meta)}/{MAX_M} car.</b>
          <span style='display:inline-block;width:60px;height:5px;background:#e5e7eb;border-radius:3px;vertical-align:middle;margin-left:4px'>
          <span style='display:block;width:{meta_bar_pct}%;height:100%;background:{meta_bar_col};border-radius:3px'></span></span>
        </span>
      </div>
    </div>"""

def detect_cannibalization(df):
    """Détecte les pages ciblant les mêmes mots-clés principaux"""
    cannibal=[]
    html_df=df[df["file_type"]=="html"].copy() if "file_type" in df.columns else df.copy()
    kw_map={}
    for _,row in html_df.iterrows():
        sample=row.get("body_text_sample","")
        if not sample: continue
        kws=extract_keywords(sample, top_n=3)
        top3=[k for k,_ in kws]
        for kw in top3:
            if kw not in kw_map: kw_map[kw]=[]
            kw_map[kw].append(row["url"])
    for kw, urls in kw_map.items():
        if len(urls)>=2:
            cannibal.append({"mot_clé":kw,"nb_pages":len(urls),
                             "pages":" | ".join(urls[:4])})
    return sorted(cannibal, key=lambda x:-x["nb_pages"])

def flesch_label(score):
    if score is None: return "N/A","#9ca3af"
    if score>=70: return "Facile","#059669"
    if score>=50: return "Standard","#f8ba07"
    if score>=30: return "Difficile","#f97316"
    return "Très difficile","#ef4444"

def extract_keywords(text, top_n=10):
    """Extrait les mots-clés les plus fréquents (filtre stopwords FR/EN)"""
    stopwords_fr={"le","la","les","de","du","des","un","une","et","en","au","aux","ce","se",
        "sa","son","ses","mon","ton","notre","votre","leur","leurs","que","qui","quoi","dont",
        "où","par","pour","sur","sous","dans","avec","sans","mais","ou","donc","car","ni","ne",
        "pas","plus","très","aussi","comme","si","il","elle","ils","elles","nous","vous","je","tu",
        "est","sont","être","avoir","a","ont","été","cette","tout","tous","toute","toutes",
        "the","and","for","are","was","with","that","this","from","have","has","not","but","its"}
    words=re.findall(r'\b[a-zA-ZÀ-ÿ]{3,}\b', text.lower())
    filtered=[w for w in words if w not in stopwords_fr]
    return Counter(filtered).most_common(top_n)

def analyze_page_keywords(r_dict, title, h1, meta):
    """Analyse densité + présence mots-clés dans title/h1/méta"""
    body=r_dict.get("body_text_sample","")
    if not body: return []
    kws=extract_keywords(body, top_n=8)
    results=[]
    for kw, count in kws:
        total_words=max(r_dict.get("word_count",1),1)
        density=round(count/total_words*100,2)
        in_title=kw.lower() in title.lower() if title else False
        in_h1=kw.lower() in h1.lower() if h1 else False
        in_meta=kw.lower() in meta.lower() if meta else False
        results.append({"mot_clé":kw,"occurrences":count,"densité_%":density,
                        "dans_title":"✅" if in_title else "❌",
                        "dans_h1":"✅" if in_h1 else "❌",
                        "dans_méta":"✅" if in_meta else "❌"})
    return results

def generate_suggestions(df, expert_mode=False, security_data=None):
    sugs=[]; total=len(df)
    if total==0: return sugs
    no_title=len(df[df["title"]==""])
    no_meta=len(df[df["meta_description"]==""])
    no_h1=len(df[df["h1"]==""]) if "h1" in df.columns else 0
    imgs_no_alt=int(df["images_no_alt"].sum()) if "images_no_alt" in df.columns else 0
    err404=len(df[df["status"]==404])
    slow=len(df[df["load_time_ms"]>3000]) if "load_time_ms" in df.columns else 0
    noindex=len(df[df["meta_noindex"]==True]) if "meta_noindex" in df.columns else 0
    non_webp=int(df["images_non_webp"].sum()) if "images_non_webp" in df.columns else 0
    empty_cta=int(df["cta_empty"].sum()) if "cta_empty" in df.columns else 0
    hard_read=len(df[df["flesch_score"].notna() & (df["flesch_score"]<30)]) if "flesch_score" in df.columns else 0

    # ── Alertes sécurité dans les priorités (injectées en premier si critiques)
    if security_data:
        if not security_data.get("https"):
            sugs.append({"p":"🔴 Critique","c":"red","t":"🔒 Site non sécurisé (HTTP)","a":"Installer un certificat SSL/TLS et forcer la redirection HTTP→HTTPS.","i":"Pénalité Google + données non chiffrées"})
        if security_data.get("insecure_forms"):
            n=len(security_data["insecure_forms"])
            sugs.append({"p":"🔴 Critique","c":"red","t":f"🔒 {n} formulaire(s) non sécurisé(s)","a":"Soumettre tous les formulaires en HTTPS - risque de vol de données (mots de passe).","i":"Risque RGPD + confiance utilisateurs"})
        if security_data.get("https") and not security_data.get("hsts"):
            sugs.append({"p":"🟠 Important","c":"yellow","t":"🔒 HSTS absent","a":"Ajouter Strict-Transport-Security: max-age=31536000; includeSubDomains","i":"Attaques de downgrade HTTP possibles"})
        if not security_data.get("csp"):
            sugs.append({"p":"🟠 Important","c":"yellow","t":"🔒 Content-Security-Policy manquante","a":"Définir une CSP pour bloquer les injections XSS et ressources non autorisées.","i":"Protection contre les scripts malveillants"})
        if security_data.get("mixed_content_pages"):
            n=len(security_data["mixed_content_pages"])
            sugs.append({"p":"🟠 Important","c":"yellow","t":f"🔒 Contenu mixte HTTP/HTTPS sur {n} page(s)","a":"Remplacer toutes les ressources HTTP par leur équivalent HTTPS.","i":"Ressources bloquées par le navigateur"})
        no_sec_ck=security_data.get("cookies_no_secure",0)
        no_ho_ck=security_data.get("cookies_no_httponly",0)
        if no_sec_ck>0 or no_ho_ck>0:
            sugs.append({"p":"🟡 Normal","c":"blue","t":f"🔒 Cookies non conformes ({no_sec_ck} sans Secure, {no_ho_ck} sans HttpOnly)","a":"Ajouter les flags Secure, HttpOnly et SameSite sur tous les cookies.","i":"Protection CSRF et XSS"})
        if not security_data.get("x_frame"):
            sugs.append({"p":"🟡 Normal","c":"blue","t":"🔒 X-Frame-Options absent","a":"Ajouter X-Frame-Options: SAMEORIGIN pour prévenir le clickjacking.","i":"Protection des utilisateurs"})

    if no_title>0: sugs.append({"p":"🔴 Critique","c":"red","t":f"{no_title} page(s) sans Title","a":"Rédiger un title unique 30-60 car. avec le mot-clé principal.","i":"Signal SEO n°1 - impact direct sur le classement"})
    if err404>0: sugs.append({"p":"🔴 Critique","c":"red","t":f"{err404} erreur(s) 404","a":"Créer des redirections 301 vers les pages remplaçantes.","i":"Perte de jus SEO et dégradation UX"})
    if noindex>0: sugs.append({"p":"🔴 Critique","c":"red","t":f"{noindex} page(s) noindex","a":"Vérifier si ces pages doivent être indexées - retirer noindex si erreur.","i":"Pages invisibles pour Google"})
    if no_meta>0: sugs.append({"p":"🟠 Important","c":"yellow","t":f"{no_meta} page(s) sans méta-description","a":"Rédiger une méta unique de 70-155 car. incitative au clic.","i":"Impact direct sur le CTR en SERP"})
    if no_h1>0: sugs.append({"p":"🟠 Important","c":"yellow","t":f"{no_h1} page(s) sans H1","a":"Ajouter un H1 unique et optimisé par page.","i":"Signal sémantique fort pour Google"})
    if imgs_no_alt>0: sugs.append({"p":"🟠 Important","c":"yellow","t":f"{imgs_no_alt} image(s) sans ALT","a":"Renseigner l'attribut ALT avec une description courte.","i":"SEO image + accessibilité WCAG"})
    if slow>0: sugs.append({"p":"🟠 Important","c":"yellow","t":f"{slow} page(s) > 3s de chargement","a":"Compresser images (WebP), activer le cache, réduire le JS bloquant.","i":"-7% de conversions par seconde supplémentaire"})
    if non_webp>0: sugs.append({"p":"🟠 Important","c":"yellow","t":f"{non_webp} image(s) non WebP","a":"Convertir en WebP - gain de 25-35% sur le poids des images.","i":"Impact direct sur LCP et PageSpeed"})
    if empty_cta>0: sugs.append({"p":"🟠 Important","c":"yellow","t":f"{empty_cta} CTA / lien(s) sans texte","a":"Ajouter un texte ou un aria-label sur chaque bouton/lien vide.","i":"Accessibilité WCAG + UX + SEO interne"})
    if hard_read>0: sugs.append({"p":"🟡 Normal","c":"blue","t":f"{hard_read} page(s) à lisibilité difficile","a":"Simplifier les phrases, réduire le jargon, viser Flesch ≥ 50.","i":"Rétention des visiteurs + taux de conversion"})
    if expert_mode:
        long_t=len(df[df["title_len"]>60]) if "title_len" in df.columns else 0
        no_og=len(df[df["og_title"]==""]) if "og_title" in df.columns else 0
        no_schema=len(df[df["has_schema"]==False]) if "has_schema" in df.columns else 0
        if long_t>0: sugs.append({"p":"🟡 Normal","c":"blue","t":f"{long_t} title(s) > 60 car.","a":"Raccourcir à 55-60 car. pour éviter la troncature en SERP.","i":"Title tronqué = moins de clics"})
        if no_og>0: sugs.append({"p":"🟡 Normal","c":"blue","t":f"{no_og} page(s) sans Open Graph","a":"Ajouter og:title, og:description, og:image sur chaque page.","i":"Prévisualisation réseaux sociaux améliorée"})
        if no_schema>0: sugs.append({"p":"🟡 Normal","c":"blue","t":f"{no_schema} page(s) sans données structurées","a":"Ajouter Schema.org (JSON-LD) pour les rich snippets Google.","i":"Rich snippets = +30% CTR en moyenne"})
    return sugs

def generate_pdf(df, domain, score, ps_data, dups, sugs, opr_data=None, security_data=None):
    try: from fpdf import FPDF
    except: return None

    def safe(text):
        """Convertit une chaîne en latin-1 safe — remplace emojis et caractères spéciaux"""
        if not text: return ""
        text = str(text)
        replacements = {
            "\U0001f534": "[CRITIQUE]", "\U0001f7e0": "[IMPORTANT]", "\U0001f7e1": "[NORMAL]", "\U0001f7e2": "[OK]",
            "\u2705": "OK", "\u274c": "NON", "\u26a0\ufe0f": "(!)", "\u26a0": "(!)", "\U0001f512": "[SEC]",
            "\U0001f333": "", "\U0001f4dd": "", "\U0001f50e": "", "\U0001f4c4": "", "\U0001f517": "",
            "\u23f1": "", "\U0001f501": "", "\u26a1": "", "\u267f": "", "\U0001f5fa": "",
            "\U0001f916": "", "\U0001f310": "", "\U0001f511": "", "\U0001f4ca": "",
            "\u2192": "->", "\u2014": "-", "\u2013": "-", "\u00ab": '"', "\u00bb": '"',
            "\u2018": "'", "\u2019": "'", "\u201c": '"', "\u201d": '"', "\u2026": "...", "\u00b7": ".",
            "\u0153": "oe", "\u0152": "OE", "\u00e6": "ae", "\u00c6": "AE",
            "\u2022": "-", "\u25cf": "-",
        }
        for k, v in replacements.items():
            text = text.replace(k, v)
        result = ""
        for ch in text:
            try:
                ch.encode("latin-1")
                result += ch
            except (UnicodeEncodeError, UnicodeDecodeError):
                result += "?"
        return result

    # ── Palette Scopo
    NAVY   = (17,24,39)
    BLUE   = (37,99,235)
    ORANGE = (37,99,235)
    YELLOW = (245,158,11)
    TEAL   = (96,165,250)
    WHITE  = (255,255,255)
    GRAY   = (107,114,128)
    LIGHT  = (248,249,250)
    GREEN  = (16,185,129)
    RED    = (239,68,68)
    domain_clean = domain.replace("https://","").replace("http://","").rstrip("/")
    date_str = datetime.datetime.now().strftime("%d/%m/%Y")
    total = len(df)
    errors = len(df[df["status"]!=200])

    class PDF(FPDF):
        def header(self):
            if self.page_no()==1: return   # page de garde : pas de header répété
            # Bande navy fine
            self.set_fill_color(*NAVY); self.rect(0,0,210,14,'F')
            # Trait orange
            self.set_fill_color(*ORANGE); self.rect(0,14,210,1.5,'F')
            # Logo texte gauche
            self.set_xy(8,3); self.set_font("Helvetica","B",9)
            self.set_text_color(*WHITE); self.cell(40,8,"SCOPO",new_x="RIGHT", new_y="TOP")
            self.set_font("Helvetica","",7); self.set_text_color(170,180,220)
            self.cell(0,8,f"Audit SEO  -  {domain_clean}  -  {date_str}",new_x="RIGHT", new_y="TOP",align="R")
            self.set_xy(0,17)

        def footer(self):
            if self.page_no()==1: return
            # Bande bas navy
            self.set_fill_color(*NAVY); self.rect(0,284,210,13,'F')
            self.set_fill_color(*ORANGE); self.rect(0,284,210,1.5,'F')
            self.set_y(286)
            self.set_font("Helvetica","",7); self.set_text_color(170,180,220)
            self.cell(0,5,
                f"Scopo - SEO Audit, simplified.  -  hello@scopo.app  -  scopo.app  -  Page {self.page_no()}",
                align="C")

        def section(self, title, col=BLUE, icon=""):
            self.set_fill_color(*col); self.set_text_color(*WHITE)
            self.set_font("Helvetica","B",10)
            self.cell(0,8,f"  {icon}  {title}" if icon else f"  {title}",new_x="LMARGIN", new_y="NEXT",fill=True)
            self.set_text_color(0,0,0); self.ln(3)

        def kv(self, label, value, val_color=None):
            self.set_font("Helvetica","",9); self.set_text_color(*GRAY)
            self.cell(72,5,label)
            self.set_font("Helvetica","B",9)
            self.set_text_color(*(val_color or NAVY))
            self.cell(0,5,str(value),new_x="LMARGIN", new_y="NEXT")

        def badge(self, x, y, w, h, text, bg, fg=WHITE):
            self.set_fill_color(*bg); self.set_text_color(*fg)
            self.set_font("Helvetica","B",8)
            self.set_xy(x,y); self.cell(w,h,text,fill=True,align="C")

        def score_gauge(self, x, y, value, label, max_w=60):
            """Dessine une mini-jauge horizontale colorée"""
            col = GREEN if value>=70 else (YELLOW if value>=40 else RED)
            # fond gris
            self.set_fill_color(229,231,235); self.rect(x,y,max_w,3,'F')
            # barre colorée
            fill_w = max_w * min(value,100)/100
            self.set_fill_color(*col); self.rect(x,y,fill_w,3,'F')
            # texte
            self.set_xy(x,y+4); self.set_font("Helvetica","B",8); self.set_text_color(*col)
            self.cell(max_w,4,f"{value}/100",align="C")
            self.set_xy(x,y+8); self.set_font("Helvetica","",7); self.set_text_color(*GRAY)
            self.cell(max_w,4,label,align="C")

    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    # ════════════════════════════════════════════════════════════
    # PAGE 1 - COUVERTURE HINSIGHT
    # ════════════════════════════════════════════════════════════
    pdf.add_page()

    # Bloc navy pleine hauteur haut
    pdf.set_fill_color(*NAVY); pdf.rect(0,0,210,120,'F')
    # Accent orange vertical gauche
    pdf.set_fill_color(*ORANGE); pdf.rect(0,0,6,120,'F')
    # Trait teal horizontal décoratif
    pdf.set_fill_color(*TEAL); pdf.rect(6,85,204,1.5,'F')

    # ── Texte couverture
    pdf.set_xy(16,22)
    pdf.set_font("Helvetica","B",9); pdf.set_text_color(*TEAL)
    pdf.cell(0,6,"SCOPO - SEO AUDIT, SIMPLIFIED.",new_x="LMARGIN", new_y="NEXT")

    pdf.set_x(16)
    pdf.set_font("Helvetica","B",28); pdf.set_text_color(*WHITE)
    pdf.cell(0,12,"Rapport d'audit SEO",new_x="LMARGIN", new_y="NEXT")

    pdf.set_x(16)
    pdf.set_font("Helvetica","",13); pdf.set_text_color(170,186,230)
    pdf.cell(0,8,safe(f"Analyse complete - {domain_clean}"),new_x="LMARGIN", new_y="NEXT")

    pdf.set_x(16); pdf.ln(4)
    pdf.set_font("Helvetica","B",10); pdf.set_text_color(*YELLOW)
    pdf.cell(0,6,safe(f"Genere le {datetime.datetime.now().strftime('%d %B %Y - %H:%M')}"),new_x="LMARGIN", new_y="NEXT")

    # ── Score global mis en avant sur la couverture
    sc_col = GREEN if score>=70 else (YELLOW if score>=40 else RED)
    sc_label = "Bon" if score>=70 else ("À améliorer" if score>=40 else "Critique")
    pdf.set_xy(140,30)
    pdf.set_font("Helvetica","B",42); pdf.set_text_color(*sc_col)
    pdf.cell(60,18,f"{score}",align="C",new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(140); pdf.set_font("Helvetica","B",11); pdf.set_text_color(*WHITE)
    pdf.cell(60,5,"/100",align="C",new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(140); pdf.set_font("Helvetica","",9); pdf.set_text_color(*sc_col)
    pdf.cell(60,6,safe(f"Score SEO - {sc_label}"),align="C",new_x="LMARGIN", new_y="NEXT")

    # Jauge couverture
    pdf.set_fill_color(30,40,100); pdf.rect(140,72,60,4,'F')
    fw = 60*min(score,100)//100
    pdf.set_fill_color(*sc_col); pdf.rect(140,72,fw,4,'F')

    # ── KPIs résumé couverture (positionnement absolu)
    kpis = [
        (str(total), "URLs crawlees"),
        (str(errors), "Erreurs"),
        (safe(f"{int(df['load_time_ms'].mean())} ms"), "Tps moyen"),
        (str(len(dups)), "Doublons"),
        (str(len(sugs)), "Preconisations"),
    ]
    col_w = 36
    x_start = 16
    y_kpi = 96
    for idx, (val, lbl) in enumerate(kpis):
        x = x_start + idx * col_w
        pdf.set_xy(x, y_kpi)
        pdf.set_font("Helvetica","B",14); pdf.set_text_color(*WHITE)
        pdf.cell(col_w, 7, safe(val), align="C")
        pdf.set_xy(x, y_kpi + 7)
        pdf.set_font("Helvetica","",7); pdf.set_text_color(140,160,220)
        pdf.cell(col_w, 4, safe(lbl), align="C")

    # ── Zone blanche bas de couverture
    pdf.set_fill_color(244,246,251); pdf.rect(0,120,210,177,'F')
    pdf.set_fill_color(*ORANGE); pdf.rect(0,120,6,177,'F')

    # Sommaire visuel
    pdf.set_xy(16,128)
    pdf.set_font("Helvetica","B",11); pdf.set_text_color(*NAVY)
    pdf.cell(0,7,"Sommaire de ce rapport",new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(16); pdf.set_fill_color(*ORANGE); pdf.rect(16,137,30,1,'F')
    pdf.ln(6)

    sommaire = [
        ("01", "Synthèse & score global",        "Indicateurs clés, score SEO et métriques de performance"),
        ("02", "Préconisations prioritaires",     "Actions classées par niveau de criticité et impact estimé"),
        ("03", "Sécurité du site",                "HTTPS, headers HTTP, cookies, formulaires et contenu mixte"),
        ("04", "Top pages à corriger",            "Les 20 URLs avec le score SEO le plus faible"),
        ("05", "À propos de Scopo",                "SEO Audit, simplified. — scopo.app"),
    ]
    for num, titre, desc in sommaire:
        y = pdf.get_y()
        # Numéro
        pdf.set_fill_color(*ORANGE); pdf.rect(16, y, 10, 8, 'F')
        pdf.set_xy(16, y); pdf.set_font("Helvetica","B",8); pdf.set_text_color(*WHITE)
        pdf.cell(10,8,num,align="C")
        # Titre
        pdf.set_xy(29, y); pdf.set_font("Helvetica","B",9); pdf.set_text_color(*NAVY)
        pdf.cell(0,4,titre,new_x="LMARGIN", new_y="NEXT")
        pdf.set_x(29); pdf.set_font("Helvetica","",8); pdf.set_text_color(*GRAY)
        pdf.cell(0,4,desc,new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    # ── Bloc Scopo bas de couverture
    pdf.set_xy(0,240)
    pdf.set_fill_color(*NAVY); pdf.rect(0,240,210,57,'F')
    pdf.set_fill_color(*BLUE); pdf.rect(0,240,210,1.5,'F')
    pdf.set_fill_color(*TEAL); pdf.rect(0,295,210,2,'F')

    pdf.set_xy(16,248)
    pdf.set_font("Helvetica","B",11); pdf.set_text_color(*WHITE)
    pdf.cell(0,6,"Cet audit a été réalisé avec Scopo",new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(16); pdf.set_font("Helvetica","",8); pdf.set_text_color(170,186,230)
    pdf.multi_cell(130,5,
        "Scopo est un outil d'audit SEO gratuit et open source. "
        "Crawl on-page, vitesse, sécurité, liens sortants, Open Graph, "
        "Schema.org et bien plus. SEO Audit, simplified.")

    # Contact à droite
    pdf.set_xy(155,248)
    pdf.set_font("Helvetica","B",8); pdf.set_text_color(*TEAL)
    pdf.cell(50,5,"CONTACT",align="C",new_x="LMARGIN", new_y="NEXT")
    for line in ["hello@scopo.app","scopo.app"]:
        pdf.set_x(155); pdf.set_font("Helvetica","",8); pdf.set_text_color(*WHITE)
        pdf.cell(50,5,line,align="C",new_x="LMARGIN", new_y="NEXT")

    # ════════════════════════════════════════════════════════════
    # PAGE 2 - SYNTHÈSE & SCORE
    # ════════════════════════════════════════════════════════════
    pdf.add_page(); pdf.set_margins(10,20,10); pdf.set_y(20)

    pdf.section("01 - Synthèse & score global", NAVY, "")

    # Scores en colonnes (SEO + PageSpeed si dispo)
    col_x = [12, 78, 144]
    scores_to_show = [
        (score, "Score SEO global", ""),
    ]
    if ps_data and "score" in ps_data:
        scores_to_show.append((ps_data["score"], "PageSpeed mobile", ""))
    if security_data and "score" in security_data:
        scores_to_show.append((security_data["score"], "Score sécurité", ""))

    y_gauge = pdf.get_y()+2
    for i, (sc, lbl, _) in enumerate(scores_to_show[:3]):
        pdf.score_gauge(col_x[i], y_gauge, sc, lbl, max_w=58)
    pdf.set_y(y_gauge+16)
    pdf.ln(2)

    # KPIs en tableau 2 colonnes
    pdf.section("Métriques clés", BLUE)
    kv_data = [
        ("URLs analysées", total),
        ("Erreurs (non-200)", errors),
        ("Score moyen/page", f"{int(df['seo_score'].mean())}/100"),
        ("Temps charg. moyen", f"{int(df['load_time_ms'].mean())} ms"),
        ("Pages sans Title", len(df[df["title"]==""]), RED if len(df[df["title"]==""])>0 else GREEN),
        ("Pages sans Méta", len(df[df["meta_description"]==""]), RED if len(df[df["meta_description"]==""])>0 else GREEN),
        ("Pages sans H1", len(df[df["h1"]==""]) if "h1" in df.columns else 0),
        ("Contenus dupliqués", len(dups), ORANGE if len(dups)>0 else GREEN),
    ]
    if opr_data and "page_rank_decimal" in opr_data:
        kv_data.append(("Open PageRank", f"{opr_data['page_rank_decimal']}/10 (rang #{opr_data.get('rank','N/A')})"))
    if ps_data and "lcp" in ps_data:
        kv_data.append(("LCP (mobile)", ps_data.get("lcp","N/A")))
        kv_data.append(("CLS (mobile)", ps_data.get("cls","N/A")))

    for i, item in enumerate(kv_data):
        label = item[0]; val = item[1]; vc = item[2] if len(item)>2 else NAVY
        pdf.kv(f"{label} :", val, vc)
    pdf.ln(4)

    # ════════════════════════════════════════════════════════════
    # PAGE 3 - PRÉCONISATIONS
    # ════════════════════════════════════════════════════════════
    pdf.add_page(); pdf.set_margins(10,20,10); pdf.set_y(20)
    pdf.section("02 - Préconisations prioritaires", ORANGE)

    prio_colors = {
        "🔴 Critique":  (RED,    (255,240,240), "CRITIQUE"),
        "🟠 Important": (ORANGE, (255,247,237), "IMPORTANT"),
        "🟡 Normal":    (YELLOW, (255,251,235), "NORMAL"),
    }

    for s in sugs[:12]:
        prio = s.get("p","")
        border_col, bg_col, prio_label = prio_colors.get(prio, (BLUE, LIGHT, "INFO"))
        y0 = pdf.get_y()
        card_h = 20
        pdf.set_fill_color(*bg_col); pdf.rect(10, y0, 190, card_h, 'F')
        pdf.set_fill_color(*border_col); pdf.rect(10, y0, 3, card_h, 'F')
        pdf.set_xy(15, y0+2)
        pdf.set_font("Helvetica","B",7); pdf.set_text_color(*border_col)
        pdf.cell(35,4,prio_label)
        pdf.set_xy(15, y0+7)
        pdf.set_font("Helvetica","B",8.5); pdf.set_text_color(*NAVY)
        pdf.cell(185,4,safe(s.get("t","")[:90]))
        pdf.set_xy(15, y0+13)
        pdf.set_font("Helvetica","",7.5); pdf.set_text_color(*GRAY)
        pdf.cell(185,4,safe(f"-> {s.get('a','')[:100]}"))
        pdf.set_y(y0+card_h+2)

    pdf.ln(2)

    # ════════════════════════════════════════════════════════════
    # PAGE 4 - SÉCURITÉ (si security_data dispo)
    # ════════════════════════════════════════════════════════════
    if security_data:
        pdf.add_page(); pdf.set_margins(10,20,10); pdf.set_y(20)
        sec_score = security_data.get("score",0)
        sec_col = GREEN if sec_score>=70 else (YELLOW if sec_score>=40 else RED)
        pdf.section("03 - Sécurité du site", NAVY)

        # Score sécurité
        pdf.score_gauge(12, pdf.get_y()+1, sec_score, "Score sécurité", max_w=60)
        pdf.set_xy(80, pdf.get_y()+1)
        pdf.set_font("Helvetica","",8); pdf.set_text_color(*GRAY)
        issues_txt = "  -  ".join(security_data.get("issues",[])[:4])
        pdf.multi_cell(120,4,safe(issues_txt or "Aucun probleme critique detecte"))
        pdf.set_y(pdf.get_y()+14); pdf.ln(2)

        # Tableau des 7 en-têtes
        pdf.section("En-têtes HTTP de sécurité", BLUE)
        headers_check = [
            ("HTTPS",                    security_data.get("https")),
            ("HSTS",                     security_data.get("hsts")),
            ("Content-Security-Policy",  security_data.get("csp")),
            ("X-Frame-Options",          security_data.get("x_frame")),
            ("X-Content-Type-Options",   security_data.get("x_content_type")),
            ("Referrer-Policy",          security_data.get("referrer_policy")),
            ("Permissions-Policy",       security_data.get("permissions_policy")),
        ]
        # En-tête tableau
        pdf.set_fill_color(*NAVY); pdf.set_text_color(*WHITE); pdf.set_font("Helvetica","B",8)
        pdf.cell(100,6,"En-tête",fill=True); pdf.cell(30,6,"Statut",fill=True,align="C")
        pdf.cell(0,6,"Impact",fill=True); pdf.ln()
        impacts = {
            "HTTPS":"Chiffrement du transit - obligatoire",
            "HSTS":"Protection downgrade HTTP",
            "Content-Security-Policy":"Injection XSS",
            "X-Frame-Options":"Clickjacking",
            "X-Content-Type-Options":"MIME sniffing",
            "Referrer-Policy":"Fuite d'URLs internes",
            "Permissions-Policy":"APIs navigateur non restreintes",
        }
        for i,(name,present) in enumerate(headers_check):
            bg = (240,253,244) if present else (255,241,242)
            pdf.set_fill_color(*bg)
            pdf.set_text_color(*NAVY); pdf.set_font("Helvetica","",8)
            pdf.cell(100,5.5,name,fill=True)
            status_txt = "OK" if present else "ABSENT"
            s_col = GREEN if present else RED
            pdf.set_text_color(*s_col); pdf.set_font("Helvetica","B",8)
            pdf.cell(30,5.5,status_txt,fill=True,align="C")
            pdf.set_text_color(*GRAY); pdf.set_font("Helvetica","",7)
            pdf.cell(0,5.5,impacts.get(name,""),fill=True)
            pdf.ln()
        pdf.ln(4)

        # Cookies résumé
        total_ck = len(security_data.get("cookies",[]))
        if total_ck>0:
            pdf.section("Cookies", BLUE)
            pdf.kv("Total cookies détectés :", total_ck)
            pdf.kv("Sans flag Secure :", security_data.get("cookies_no_secure",0),
                   RED if security_data.get("cookies_no_secure",0)>0 else GREEN)
            pdf.kv("Sans HttpOnly :", security_data.get("cookies_no_httponly",0),
                   RED if security_data.get("cookies_no_httponly",0)>0 else GREEN)
            pdf.kv("Sans SameSite :", security_data.get("cookies_no_samesite",0),
                   ORANGE if security_data.get("cookies_no_samesite",0)>0 else GREEN)
            pdf.ln(3)

        # Recommandations sécurité
        recs = security_data.get("recommendations",[])
        if recs:
            pdf.section("Recommandations sécurité", ORANGE)
            for rec in recs[:6]:
                p = rec.get("p","")
                bc = RED if "Critique" in p else (ORANGE if "Important" in p else YELLOW)
                bg = (255,240,240) if "Critique" in p else ((255,247,237) if "Important" in p else (255,251,235))
                lbl = "CRITIQUE" if "Critique" in p else ("IMPORTANT" if "Important" in p else "NORMAL")
                y0 = pdf.get_y()
                pdf.set_fill_color(*bg); pdf.rect(10,y0,190,18,'F')
                pdf.set_fill_color(*bc); pdf.rect(10,y0,3,18,'F')
                pdf.set_xy(15,y0+2); pdf.set_font("Helvetica","B",7); pdf.set_text_color(*bc)
                pdf.cell(35,4,lbl)
                pdf.set_xy(15,y0+7); pdf.set_font("Helvetica","B",8); pdf.set_text_color(*NAVY)
                pdf.cell(185,4,safe(rec.get("t","")[:80]))
                pdf.set_xy(15,y0+12); pdf.set_font("Helvetica","",7); pdf.set_text_color(*GRAY)
                pdf.cell(185,4,safe(f"-> {rec.get('a','')[:100]}"))
                pdf.set_y(y0+20)

    # ════════════════════════════════════════════════════════════
    # PAGE 5 - TOP PAGES À CORRIGER
    # ════════════════════════════════════════════════════════════
    pdf.add_page(); pdf.set_margins(10,20,10); pdf.set_y(20)
    section_num = "04" if security_data else "03"
    pdf.section(f"{section_num} - Top 20 pages à corriger", NAVY)

    df_s = df.sort_values("seo_score").head(20)
    # En-tête tableau
    pdf.set_fill_color(*NAVY); pdf.set_text_color(*WHITE); pdf.set_font("Helvetica","B",8)
    for col_name, col_w in [("URL",95),("Statut",15),("Score",15),("Problèmes principaux",65)]:
        pdf.cell(col_w,6,col_name,fill=True,align="C")
    pdf.ln()

    pdf.set_font("Helvetica","",7.5)
    for i,(_,row) in enumerate(df_s.iterrows()):
        bg = LIGHT if i%2==0 else WHITE
        pdf.set_fill_color(*bg)
        sc_p = row.get("seo_score",0)
        sc_c2 = GREEN if sc_p>=70 else (YELLOW if sc_p>=40 else RED)
        pdf.set_text_color(*NAVY)
        url_short = safe(row["url"].replace("https://","").replace("http://","")[:55])
        pdf.cell(95,5,url_short,fill=True)
        status = row.get("status",0)
        s_col = GREEN if status==200 else RED
        pdf.set_text_color(*s_col); pdf.cell(15,5,str(status),fill=True,align="C")
        pdf.set_text_color(*sc_c2); pdf.cell(15,5,str(sc_p),fill=True,align="C")
        issues = row.get("issues","")
        issues_str = " - ".join(issues[:3]) if isinstance(issues,list) else str(issues)[:45]
        pdf.set_text_color(*GRAY); pdf.cell(65,5,safe(issues_str[:45]),fill=True)
        pdf.ln()
    pdf.ln(6)

    # ════════════════════════════════════════════════════════════
    # DERNIÈRE PAGE - PUB HINSIGHT
    # ════════════════════════════════════════════════════════════
    pdf.add_page()
    # Fond navy plein
    pdf.set_fill_color(*NAVY); pdf.rect(0,0,210,297,'F')
    # Accent orange vertical gauche
    pdf.set_fill_color(*ORANGE); pdf.rect(0,0,6,297,'F')
    # Lignes décoratives teal
    pdf.set_fill_color(*TEAL); pdf.rect(6,100,204,1,'F')
    pdf.set_fill_color(*TEAL); pdf.rect(6,230,204,1,'F')
    pdf.set_fill_color(*TEAL); pdf.rect(6,60,204,0.5,'F')

    section_num_final = "05" if security_data else "04"
    pdf.set_xy(16,30)
    pdf.set_font("Helvetica","B",8); pdf.set_text_color(*TEAL)
    pdf.cell(0,5,safe(f"{section_num_final} - A PROPOS"),new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(16); pdf.set_font("Helvetica","B",26); pdf.set_text_color(*WHITE)
    pdf.cell(0,12,"Scopo",new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(16); pdf.set_font("Helvetica","",13); pdf.set_text_color(170,186,230)
    pdf.cell(0,7,"SEO Audit, simplified.",new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(16); pdf.set_fill_color(*BLUE); pdf.rect(16,60,80,1,'F')

    pdf.set_xy(16,70)
    pdf.set_font("Helvetica","",9); pdf.set_text_color(200,210,240)
    pdf.multi_cell(178,5.5,
        "Scopo est un outil d'audit SEO gratuit qui aide les équipes "
        "marketing et les professionnels du web à analyser et optimiser "
        "leur référencement naturel : crawl, on-page, vitesse, sécurité, "
        "accessibilité et bien plus.")

    # ── Fonctionnalités en 3 blocs
    services = [
        ("Crawl & On-page", BLUE,
         "Crawl complet - Title/H1/Meta - Canonical - Noindex - "
         "Score SEO par page - Arborescence"),
        ("Vitesse & Sécurité", TEAL,
         "Core Web Vitals - PageSpeed - HTTPS - Headers HTTP - "
         "Cookies - Contenu mixte"),
        ("Analyse avancée", YELLOW,
         "Mots-clés - Lisibilité Flesch - Open Graph - Schema.org - "
         "Maillage interne - Doublons"),
    ]
    bw = 57; bx = 16
    for title, col, desc in services:
        by = 110
        pdf.set_fill_color(*col); pdf.rect(bx, by, bw, 1.5, 'F')
        pdf.set_xy(bx, by+5); pdf.set_font("Helvetica","B",9); pdf.set_text_color(*col)
        pdf.cell(bw,5,title,new_x="LMARGIN", new_y="NEXT")
        pdf.set_x(bx); pdf.set_font("Helvetica","",7.5); pdf.set_text_color(200,210,240)
        pdf.multi_cell(bw-2, 4.5, desc)
        bx += bw + 4

    # ── Tagline
    pdf.set_xy(16,180)
    pdf.set_font("Helvetica","B",10); pdf.set_text_color(*TEAL)
    pdf.cell(0,5,"Exports inclus",new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(16); pdf.set_font("Helvetica","B",16); pdf.set_text_color(*WHITE)
    pdf.cell(0,8,"Excel  ·  PDF  ·  JSON",new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(16); pdf.set_font("Helvetica","",9); pdf.set_text_color(170,186,230)
    pdf.multi_cell(130,5,
        "Tous les rapports sont exportables en Excel multi-onglets, "
        "en PDF professionnel et en JSON pour intégration dans vos outils. "
        "Gratuit, open source, déployable sur Streamlit Cloud.")

    # ── CTA contact
    pdf.set_xy(16,240)
    pdf.set_fill_color(*BLUE); pdf.rect(16,240,80,30,'F')
    pdf.set_fill_color(*TEAL); pdf.rect(104,240,102,30,'F')

    pdf.set_xy(16,245); pdf.set_font("Helvetica","B",8); pdf.set_text_color(*WHITE)
    pdf.cell(80,5,"CONTACT",align="C",new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(16); pdf.set_font("Helvetica","",8); pdf.set_text_color(*WHITE)
    pdf.cell(80,5,"hello@scopo.app",align="C",new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(16); pdf.cell(80,5,"scopo.app",align="C")

    pdf.set_xy(104,245); pdf.set_font("Helvetica","B",8); pdf.set_text_color(*NAVY)
    pdf.cell(102,5,"RAPPORT GENERE PAR SCOPO",align="C",new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(104); pdf.set_font("Helvetica","",7.5); pdf.set_text_color(*NAVY)
    pdf.multi_cell(100,4.5,
        "Scopo - SEO Audit, simplified. "
        "Outil gratuit et open source. "
        "Rapport confidentiel - Usage exclusif du destinataire.")

    return bytes(pdf.output())

def to_excel(df, outbound_df=None, files_df=None, security_data=None):
    out=BytesIO()
    df_e=df.drop(columns=["internal_links_list","body_text_sample","body_text_hash","outbound_links_detail","outbound_domains"],errors="ignore")
    with pd.ExcelWriter(out,engine="openpyxl") as w:
        df_e.to_excel(w,sheet_name="Audit SEO",index=False)
        if outbound_df is not None and not outbound_df.empty:
            outbound_df.to_excel(w,sheet_name="Liens sortants",index=False)
        if files_df is not None and not files_df.empty:
            files_df.to_excel(w,sheet_name="Fichiers",index=False)

        # ── Onglet Sécurité
        if security_data:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Couleurs charte Scopo
            NAVY_FILL   = PatternFill("solid", fgColor="111827")
            ORANGE_FILL = PatternFill("solid", fgColor="2563EB")
            GREEN_FILL  = PatternFill("solid", fgColor="D1FAE5")
            RED_FILL    = PatternFill("solid", fgColor="FEE2E2")
            YELLOW_FILL = PatternFill("solid", fgColor="FEF9C3")
            GRAY_FILL   = PatternFill("solid", fgColor="F3F4F6")
            WHITE_FONT  = Font(color="FFFFFF", bold=True, size=11)
            NAVY_FONT   = Font(color="111827", bold=True, size=10)
            DARK_FONT   = Font(color="111827", size=10)
            BOLD_FONT   = Font(bold=True, size=10)
            thin = Side(style="thin", color="E5E7EB")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

            ws = w.book.create_sheet("🔒 Sécurité")

            # ── Titre principal
            ws.merge_cells("A1:F1")
            ws["A1"] = "Rapport Sécurité - Scopo"
            ws["A1"].fill = NAVY_FILL
            ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
            ws["A1"].alignment = center
            ws.row_dimensions[1].height = 30

            # ── Score sécurité
            score = security_data.get("score", 0)
            score_fill = PatternFill("solid", fgColor="D1FAE5") if score>=70 else (PatternFill("solid", fgColor="FEF9C3") if score>=40 else PatternFill("solid", fgColor="FEE2E2"))
            score_font = Font(color="065F46" if score>=70 else ("92400E" if score>=40 else "991B1B"), bold=True, size=20)
            ws.merge_cells("A2:C2")
            ws["A2"] = f"Score sécurité : {score}/100"
            ws["A2"].fill = score_fill
            ws["A2"].font = score_font
            ws["A2"].alignment = center
            ws.merge_cells("D2:F2")
            label = "✅ Bon" if score>=70 else ("⚠️ À améliorer" if score>=40 else "🔴 Critique")
            ws["D2"] = label
            ws["D2"].fill = score_fill
            ws["D2"].font = Font(color="065F46" if score>=70 else ("92400E" if score>=40 else "991B1B"), bold=True, size=12)
            ws["D2"].alignment = center
            ws.row_dimensions[2].height = 34

            row = 4

            # ── En-têtes de sécurité
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = "En-têtes HTTP de sécurité"
            ws[f"A{row}"].fill = ORANGE_FILL
            ws[f"A{row}"].font = WHITE_FONT
            ws[f"A{row}"].alignment = center
            ws.row_dimensions[row].height = 22
            row += 1

            headers_info = [
                ("HTTPS",                    security_data.get("https"),       "Chiffrement du transit - obligatoire pour Google",        ""),
                ("HSTS",                     security_data.get("hsts"),        "Force HTTPS - évite les attaques de downgrade",            security_data.get("hsts_value","")),
                ("Content-Security-Policy",  security_data.get("csp"),         "Bloque les injections XSS et ressources non autorisées",   (security_data.get("csp_value","") or "")[:60]),
                ("X-Frame-Options",          security_data.get("x_frame"),     "Prévient le clickjacking",                                 security_data.get("x_frame_value","")),
                ("X-Content-Type-Options",   security_data.get("x_content_type"), "Bloque le MIME sniffing",                              "nosniff" if security_data.get("x_content_type") else ""),
                ("Referrer-Policy",          security_data.get("referrer_policy"), "Contrôle la fuite des URLs internes",                 security_data.get("referrer_value","")),
                ("Permissions-Policy",       security_data.get("permissions_policy"), "Restreint l'accès aux APIs navigateur (caméra, micro…)", ""),
            ]

            # En-tête colonne
            for c, label in enumerate(["En-tête", "Statut", "Rôle", "Valeur détectée"], 1):
                cell = ws.cell(row=row, column=c, value=label)
                cell.fill = PatternFill("solid", fgColor="E0E7FF")
                cell.font = NAVY_FONT
                cell.alignment = center
                cell.border = border
            ws.row_dimensions[row].height = 18
            row += 1

            for name, present, role, value in headers_info:
                ws.cell(row=row, column=1, value=name).font = BOLD_FONT
                status_cell = ws.cell(row=row, column=2, value="✅ Présent" if present else "❌ Absent")
                status_cell.fill = GREEN_FILL if present else RED_FILL
                status_cell.font = Font(color="065F46" if present else "991B1B", bold=True, size=10)
                status_cell.alignment = center
                ws.cell(row=row, column=3, value=role).font = DARK_FONT
                ws.cell(row=row, column=4, value=value or "-").font = Font(color="6B7280", size=9)
                for c in range(1, 5):
                    ws.cell(row=row, column=c).border = border
                    ws.cell(row=row, column=c).alignment = left
                ws.row_dimensions[row].height = 18
                row += 1

            row += 1

            # ── Cookies
            cookies = security_data.get("cookies", [])
            if cookies:
                ws.merge_cells(f"A{row}:F{row}")
                ws[f"A{row}"] = f"Cookies ({len(cookies)} détectés)"
                ws[f"A{row}"].fill = ORANGE_FILL
                ws[f"A{row}"].font = WHITE_FONT
                ws[f"A{row}"].alignment = center
                ws.row_dimensions[row].height = 22
                row += 1

                # KPIs cookies
                ws.cell(row=row, column=1, value="Sans flag Secure").font = BOLD_FONT
                ws.cell(row=row, column=2, value=security_data.get("cookies_no_secure",0))
                ws.cell(row=row, column=3, value="Sans HttpOnly").font = BOLD_FONT
                ws.cell(row=row, column=4, value=security_data.get("cookies_no_httponly",0))
                ws.cell(row=row, column=5, value="Sans SameSite").font = BOLD_FONT
                ws.cell(row=row, column=6, value=security_data.get("cookies_no_samesite",0))
                for c in range(1,7):
                    ws.cell(row=row, column=c).alignment = center
                    ws.cell(row=row, column=c).border = border
                ws.row_dimensions[row].height = 18
                row += 1

                # Tableau cookies
                for c, lbl in enumerate(["Nom", "Domaine", "Secure", "HttpOnly", "SameSite", "Expiration"], 1):
                    cell = ws.cell(row=row, column=c, value=lbl)
                    cell.fill = PatternFill("solid", fgColor="E0E7FF")
                    cell.font = NAVY_FONT
                    cell.alignment = center
                    cell.border = border
                ws.row_dimensions[row].height = 18
                row += 1

                for ck in cookies[:50]:
                    vals = [
                        ck.get("name",""), ck.get("domain",""),
                        "✅" if ck.get("secure") else "❌",
                        "✅" if ck.get("httponly") else "❌",
                        ck.get("samesite","") or "❌",
                        str(ck.get("expires","")) or "Session"
                    ]
                    for c, v in enumerate(vals, 1):
                        cell = ws.cell(row=row, column=c, value=v)
                        cell.font = DARK_FONT
                        cell.alignment = center
                        cell.border = border
                        if c in (3,4,5):
                            cell.fill = GREEN_FILL if v=="✅" else (RED_FILL if v=="❌" else GRAY_FILL)
                    ws.row_dimensions[row].height = 16
                    row += 1

                row += 1

            # ── Contenu mixte
            mixed = security_data.get("mixed_content_pages", [])
            if mixed:
                ws.merge_cells(f"A{row}:F{row}")
                ws[f"A{row}"] = f"Contenu mixte HTTP/HTTPS - {len(mixed)} page(s) concernée(s)"
                ws[f"A{row}"].fill = PatternFill("solid", fgColor="FEF3C7")
                ws[f"A{row}"].font = Font(color="92400E", bold=True, size=11)
                ws[f"A{row}"].alignment = center
                ws.row_dimensions[row].height = 22
                row += 1
                for mp in mixed:
                    ws.merge_cells(f"A{row}:B{row}")
                    ws[f"A{row}"] = mp.get("url","")
                    ws[f"A{row}"].font = Font(bold=True, size=9)
                    ws.cell(row=row, column=3, value=f"{mp.get('count',0)} ressource(s)")
                    for res in mp.get("ressources",[])[:3]:
                        ws.merge_cells(f"D{row}:F{row}")
                        ws[f"D{row}"] = res
                        ws[f"D{row}"].font = Font(color="6B7280", size=8)
                        row += 1
                    row += 1
                row += 1

            # ── Formulaires non sécurisés
            insecure = security_data.get("insecure_forms", [])
            if insecure:
                ws.merge_cells(f"A{row}:F{row}")
                ws[f"A{row}"] = f"Formulaires non sécurisés - {len(insecure)}"
                ws[f"A{row}"].fill = RED_FILL
                ws[f"A{row}"].font = Font(color="991B1B", bold=True, size=11)
                ws[f"A{row}"].alignment = center
                ws.row_dimensions[row].height = 22
                row += 1
                for c, lbl in enumerate(["Page", "Action formulaire", "Champ mot de passe"], 1):
                    cell = ws.cell(row=row, column=c, value=lbl)
                    cell.fill = PatternFill("solid", fgColor="E0E7FF")
                    cell.font = NAVY_FONT
                    cell.border = border
                row += 1
                for f in insecure:
                    ws.cell(row=row, column=1, value=f.get("page","")).font = DARK_FONT
                    ws.cell(row=row, column=2, value=f.get("action","")).font = DARK_FONT
                    ws.cell(row=row, column=3, value="🔐 Oui" if f.get("has_password") else "Non").font = DARK_FONT
                    for c in range(1,4):
                        ws.cell(row=row, column=c).border = border
                    row += 1

            # ── Recommandations
            row += 1
            recs = security_data.get("recommendations", [])
            if recs:
                ws.merge_cells(f"A{row}:F{row}")
                ws[f"A{row}"] = "Recommandations"
                ws[f"A{row}"].fill = NAVY_FILL
                ws[f"A{row}"].font = WHITE_FONT
                ws[f"A{row}"].alignment = center
                ws.row_dimensions[row].height = 22
                row += 1
                for c, lbl in enumerate(["Priorité", "Titre", "Action", "Impact"], 1):
                    cell = ws.cell(row=row, column=c, value=lbl)
                    cell.fill = PatternFill("solid", fgColor="E0E7FF")
                    cell.font = NAVY_FONT
                    cell.border = border
                row += 1
                for rec in recs:
                    p = rec.get("p","")
                    prio_fill = RED_FILL if "Critique" in p else (PatternFill("solid", fgColor="FFEDD5") if "Important" in p else YELLOW_FILL)
                    ws.cell(row=row, column=1, value=p).fill = prio_fill
                    ws.cell(row=row, column=2, value=rec.get("t","")).font = BOLD_FONT
                    ws.merge_cells(f"C{row}:D{row}")
                    ws.cell(row=row, column=3, value=rec.get("a","")).font = DARK_FONT
                    ws.merge_cells(f"E{row}:F{row}")
                    ws.cell(row=row, column=5, value=rec.get("impact","")).font = Font(color="6B7280", size=9)
                    for c in range(1, 7):
                        ws.cell(row=row, column=c).border = border
                        ws.cell(row=row, column=c).alignment = left
                    ws.row_dimensions[row].height = 20
                    row += 1

            # Largeurs colonnes
            col_widths = [35, 18, 45, 35, 20, 22]
            for i, w_val in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w_val

        # Largeurs colonnes onglets standards
        for sn, sheet in w.sheets.items():
            if sn == "🔒 Sécurité": continue
            for col in sheet.iter_cols(min_row=1, max_row=1):
                sheet.column_dimensions[col[0].column_letter].width = 18
    return out.getvalue()

# ── SESSION STATE
for k,v in [("crawl_done",False),("results",[]),("ps",{}),("ps_desktop",{}),("opr",{}),("domain",""),("history",[]),("robots_data",{}),("sitemap_data",{}),("cwv_history",[]),("security_data",{})]:
    if k not in st.session_state: st.session_state[k]=v

# ── HEADER
st.markdown("""<div class="app-header">
  <div>
    <h1 style="color:#111827 !important;margin:0;font-size:1.5rem;font-weight:800;letter-spacing:-.02em;font-family:'Inter',sans-serif">Scopo<span style="color:#2563EB">·</span></h1>
    <div style="color:#6B7280;margin:3px 0 0;font-size:.85rem">SEO Audit, simplified.</div>
  </div>
</div>""",unsafe_allow_html=True)

# ── SIDEBAR
with st.sidebar:
    st.markdown("### ⚙️ Configuration")
    with st.form("audit_form", clear_on_submit=False):
        url_input=st.text_input("URL de départ", placeholder="https://exemple.fr",
            help="Appuyez sur Entrée ou cliquez sur Lancer")

        # Slider URLs avec option illimitée
        max_pages_opt = st.select_slider("URLs max",
            options=[50, 100, 200, 500, 1000, 2000, 5000, 99999],
            value=100,
            format_func=lambda x: "Illimité ⚠️" if x==99999 else str(x))
        max_pages = max_pages_opt
        if max_pages == 99999:
            st.markdown("<div style='color:#f59e0b;font-size:.75rem;padding:2px 0'>⚠️ Peut prendre plusieurs heures sur un grand site</div>", unsafe_allow_html=True)

        delay=st.slider("Délai (s)",0.0,3.0,0.5,0.1)

        st.markdown("---")
        st.markdown("### 🎓 Mode")
        expert_mode=st.checkbox("⚡ Mode Expert",value=False)

        st.markdown("---")
        st.markdown("### 📄 Filtre par type de fichier")
        crawl_html=st.checkbox("Pages HTML",value=True)
        crawl_pdf=st.checkbox("PDF",value=True)
        crawl_images=st.checkbox("Images",value=False)
        crawl_docs=st.checkbox("Documents Office",value=True)
        crawl_media=st.checkbox("Médias (vidéo/audio)",value=False)
        crawl_other=st.checkbox("Autres fichiers",value=False)

        st.markdown("---")
        st.markdown("### 🔑 APIs gratuites")
        api_psi=st.text_input("Clé PageSpeed Google",type="password",help="console.cloud.google.com - gratuit")
        api_opr=st.text_input("Clé Open PageRank",type="password",help="openpagerank.com - gratuit 100 req/jour")

        st.markdown("---")
        st.markdown("### 🔎 Exclusions")
        exclude_ext=st.text_input("Extensions à ignorer",value=".css,.js,.xml")
        exclude_pat=st.text_input("Patterns URL à ignorer",value="/wp-admin,/feed")

        st.markdown("---")
        launch=st.form_submit_button("🚀 Lancer l'audit", type="primary", use_container_width=True)

# ── CRAWL
if launch:
    if not url_input:
        st.error("⚠️ Veuillez entrer une URL avant de lancer l'audit."); st.stop()
    # Auto-corriger l'URL si pas de schéma
    url_input = url_input.strip()
    if not url_input.startswith("http"):
        url_input = "https://" + url_input

    base_url=get_domain(url_input)
    start_url=url_input.rstrip("/")
    excl_ext=[e.strip() for e in exclude_ext.split(",") if e.strip()]
    excl_pat=[p.strip() for p in exclude_pat.split(",") if p.strip()]

    # Types autorisés
    allowed_types=set()
    if crawl_html: allowed_types.add("html")
    if crawl_pdf: allowed_types.add("pdf")
    if crawl_images: allowed_types.add("image")
    if crawl_docs: allowed_types.add("doc")
    if crawl_media: allowed_types.add("media")
    if crawl_other: allowed_types.update(["zip","code"])

    progress_bar=st.progress(0)
    status_text=st.empty()
    c1,c2,c3,c4=st.columns(4)
    m1=c1.empty();m2=c2.empty();m3=c3.empty();m4=c4.empty()

    visited=set(); queue=deque([start_url]); results=[]; err_n=0

    while queue and len(results)<max_pages:
        url=normalize_url(queue.popleft())
        if url in visited: continue
        if should_exclude(url,excl_ext,excl_pat): continue
        if not is_internal(url,base_url): continue
        visited.add(url)

        ftype=get_file_type(url)
        if ftype not in allowed_types: continue

        progress_bar.progress(min(len(results)/max_pages,1.0))
        status_text.markdown(f"**Crawl…** `{url}`  -  `{len(results)}` pages / `{len(queue)}` en attente")

        if ftype=="html":
            resp,elapsed=fetch(url)
            result=analyze_page(url,resp,elapsed)
            for link in result.get("internal_links_list",[]):
                ln=normalize_url(link)
                if ln not in visited: queue.append(ln)
        else:
            # Pour les fichiers non-HTML : HEAD request
            status_code,size_bytes=fetch_head(url)
            result={"url":url,"status":status_code,"file_type":ftype,
                    "file_size_kb":round(size_bytes/1024,1) if size_bytes else 0,
                    "title":"","h1":"","meta_description":"","seo_score":0,
                    "load_time_ms":0,"issues":[f"Fichier {ftype.upper()}"],
                    "redirected_to":"","links_internal":0,"links_external":0,
                    "outbound_domains":[],"outbound_links_detail":[],
                    "word_count":0,"images_total":0,"images_no_alt":0,
                    "title_len":0,"h1_count":0,"meta_desc_len":0,
                    "meta_noindex":False,"canonical":"","canonical_self":True,
                    "og_title":"","og_description":"","og_image":"",
                    "has_schema":False,"hreflang_count":0,
                    "h2_count":0,"h3_count":0,"depth":0,
                    "body_text_hash":"","body_text_sample":"",
                    "internal_links_list":[]}

        results.append(result)
        if result.get("status",200) not in (200,): err_n+=1
        avg_sc=int(sum(r.get("seo_score",0) for r in results)/len(results))
        m1.metric("Crawlées",len(results)); m2.metric("Erreurs",err_n)
        m3.metric("Score moyen",f"{avg_sc}/100"); m4.metric("File",len(queue))
        time.sleep(delay)

    progress_bar.progress(1.0)
    status_text.success(f"✅ {len(results)} ressources analysées")

    # ── Robots.txt + Sitemap
    with st.spinner("🤖 Analyse robots.txt & sitemap…"):
        robots_data=analyze_robots(base_url)
        sitemap_data=analyze_sitemap(base_url, robots_data)

    # ── Audit Sécurité & Cookies
    with st.spinner("🔒 Audit sécurité & cookies…"):
        security_data=analyze_security(base_url, results)

    # APIs gratuites
    ps_result={}; ps_desktop={}
    if api_psi:
        with st.spinner("⚡ PageSpeed mobile…"):
            ps_result=get_pagespeed(start_url,api_psi,"mobile")
        with st.spinner("🖥 PageSpeed desktop…"):
            ps_desktop=get_pagespeed(start_url,api_psi,"desktop")

    opr_result={}
    if api_opr:
        with st.spinner("🔗 Open PageRank…"):
            opr_result=get_open_pagerank(base_url,api_opr)
            if "page_rank_decimal" in opr_result:
                st.success(f"PageRank : {opr_result['page_rank_decimal']}/10 - Rang #{opr_result.get('rank','N/A')}")

    df_new=pd.DataFrame(results)
    entry={"date":datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
           "domain":base_url,"total":len(results),
           "score":site_score(df_new),"errors":err_n,
           "no_title":len(df_new[df_new["title"]==""]) if "title" in df_new.columns else 0,
           "no_meta":len(df_new[df_new["meta_description"]==""]) if "meta_description" in df_new.columns else 0}
    hist=st.session_state["history"]; hist.append(entry)
    if len(hist)>10: hist=hist[-10:]

    # Historique CWV
    cwv_entry={"date":datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),"domain":base_url}
    if ps_result and "score" in ps_result:
        cwv_entry["mobile_score"]=ps_result["score"]
        cwv_entry["mobile_lcp"]=ps_result.get("lcp","N/A")
        cwv_entry["mobile_cls"]=ps_result.get("cls","N/A")
        cwv_entry["accessibility_score"]=ps_result.get("accessibility_score")
        cwv_entry["render_blocking"]=ps_result.get("render_blocking_count",0)
    if ps_desktop and "score" in ps_desktop:
        cwv_entry["desktop_score"]=ps_desktop["score"]
        cwv_entry["desktop_lcp"]=ps_desktop.get("lcp","N/A")
    cwv_hist=st.session_state["cwv_history"]; cwv_hist.append(cwv_entry)
    if len(cwv_hist)>10: cwv_hist=cwv_hist[-10:]

    st.session_state.update({"results":results,"ps":ps_result,"ps_desktop":ps_desktop,
                              "opr":opr_result,"domain":base_url,"crawl_done":True,
                              "history":hist,"robots_data":robots_data,"sitemap_data":sitemap_data,
                              "cwv_history":cwv_hist,"security_data":security_data})

# ── RÉSULTATS
if st.session_state["crawl_done"] and st.session_state["results"]:
    results=st.session_state["results"]
    ps_result=st.session_state["ps"]
    ps_desktop=st.session_state.get("ps_desktop",{})
    opr_result=st.session_state["opr"]
    robots_data=st.session_state.get("robots_data",{})
    sitemap_data=st.session_state.get("sitemap_data",{})
    security_data=st.session_state.get("security_data",{})
    cwv_history=st.session_state.get("cwv_history",[])
    domain=st.session_state["domain"]
    df=pd.DataFrame(results)

    total=len(df)
    errors=len(df[df["status"]!=200])
    global_score=site_score(df)
    dups=detect_duplicates(results)
    sugs=generate_suggestions(df,expert_mode,security_data=security_data)

    # Préparer les dataframes spéciaux
    # Liens sortants
    all_outbound=[]
    for r in results:
        for lnk in r.get("outbound_links_detail",[]):
            all_outbound.append({"page_source":r["url"],"url_cible":lnk["url"],
                                  "domaine":lnk["domain"],"ancre":lnk["anchor"]})
    df_outbound=pd.DataFrame(all_outbound) if all_outbound else pd.DataFrame()

    # Fichiers non-HTML
    df_files=df[df["file_type"]!="html"][["url","file_type","file_size_kb","status"]].copy() if "file_type" in df.columns else pd.DataFrame()

    st.markdown("---")

    # SCORE + MÉTRIQUES
    sc_color="#059669" if global_score>=70 else ("#f8ba07" if global_score>=40 else "#ef4444")
    sc_label="Excellent" if global_score>=85 else ("Bon" if global_score>=70 else ("À améliorer" if global_score>=40 else "Critique"))

    no_title_n = len(df[df["title"]==""]) if "title" in df.columns else 0
    no_meta_n  = len(df[df["meta_description"]==""]) if "meta_description" in df.columns else 0

    col_score, col_kpis = st.columns([1, 3])

    with col_score:
        st.markdown(f"""
        <div class="score-wrap">
          <div class="score-big" style="color:{sc_color}">{global_score}</div>
          <div class="score-sub">/ 100</div>
          <div class="score-lbl" style="color:{sc_color}">{sc_label}</div>
          <div style="font-size:.78rem;color:#9ca3af;margin-top:2px">Score global du site</div>
          <div class="score-bar"><div class="score-bar-fill" style="width:{global_score}%;background:{sc_color}"></div></div>
        </div>""", unsafe_allow_html=True)

    with col_kpis:
        opr_html = ""
        if opr_result and "page_rank_decimal" in opr_result:
            opr_html = f"""
            <div class="kpi-card purple" style="grid-column:span 2">
              <div style="display:flex;align-items:baseline;gap:8px">
                <div class="kpi-val">{opr_result['page_rank_decimal']}</div>
                <div style="font-size:1rem;color:#7c3aed;font-weight:600">/10</div>
              </div>
              <div class="kpi-lbl">Open PageRank - Rang #{opr_result.get('rank','N/A')}</div>
            </div>"""

        ps_html = ""
        if ps_result and "score" in ps_result:
            ps_sc = ps_result['score']
            ps_col = "#059669" if ps_sc>=70 else ("#f8ba07" if ps_sc>=40 else "#ef4444")
            ps_html = f"""
            <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-top:12px">
              <div class="kpi-card blue"><div class="kpi-val" style="color:{ps_col}">{ps_sc}</div><div class="kpi-lbl">PageSpeed /100</div></div>
              <div class="kpi-card blue"><div class="kpi-val" style="font-size:1.4rem">{ps_result.get('lcp','N/A')}</div><div class="kpi-lbl">LCP</div></div>
              <div class="kpi-card blue"><div class="kpi-val" style="font-size:1.4rem">{ps_result.get('cls','N/A')}</div><div class="kpi-lbl">CLS</div></div>
              <div class="kpi-card blue"><div class="kpi-val" style="font-size:1.4rem">{ps_result.get('tbt','N/A')}</div><div class="kpi-lbl">TBT</div></div>
            </div>"""

        st.markdown(f"""
        <div class="kpi-grid">
          <div class="kpi-card orange">
            <div class="kpi-val">{total}</div>
            <div class="kpi-lbl">URLs analysées</div>
          </div>
          <div class="kpi-card {'red' if errors>0 else 'green'}">
            <div class="kpi-val" style="color:{'#ef4444' if errors>0 else '#059669'}">{errors}</div>
            <div class="kpi-lbl">Erreurs HTTP</div>
          </div>
          <div class="kpi-card {'red' if no_title_n>0 else 'green'}">
            <div class="kpi-val" style="color:{'#ef4444' if no_title_n>0 else '#059669'}">{no_title_n}</div>
            <div class="kpi-lbl">Sans Title</div>
          </div>
          <div class="kpi-card {'red' if no_meta_n>0 else 'green'}">
            <div class="kpi-val" style="color:{'#ef4444' if no_meta_n>0 else '#059669'}">{no_meta_n}</div>
            <div class="kpi-lbl">Sans Méta</div>
          </div>
          <div class="kpi-card teal">
            <div class="kpi-val">{len(dups)}</div>
            <div class="kpi-lbl">Duplicats</div>
          </div>
          {opr_html}
        </div>
        {ps_html}
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 💡 Suggestions de corrections")
    if not sugs:
        st.success("🎉 Aucun problème majeur - bon travail !")
    else:
        bmap={"red":"#ef4444","yellow":"#f8ba07","blue":"#2563EB","green":"#10B981"}
        bgmap={"red":"rgba(239,68,68,.08)","yellow":"rgba(245,158,11,.08)","blue":"rgba(37,99,235,.08)","green":"rgba(16,185,129,.08)"}
        chipmap={"red":"rgba(239,68,68,.12);color:#EF4444","yellow":"rgba(245,158,11,.12);color:#F59E0B",
                 "blue":"rgba(37,99,235,.12);color:#2563EB","green":"rgba(16,185,129,.12);color:#10B981"}
        cols_sug = st.columns(min(len(sugs), 2))
        for idx, s in enumerate(sugs):
            b=bmap.get(s["c"],"#6b7280")
            bg=bgmap.get(s["c"],"#f9fafb")
            chip=chipmap.get(s["c"],"#f3f4f6;color:#374151")
            with cols_sug[idx % 2]:
                st.markdown(f"""<div style="background:{bg};border-left:4px solid {b};border-radius:10px;
                    padding:.9rem 1.1rem;margin-bottom:.8rem;box-shadow:0 1px 4px rgba(0,0,0,.05)">
                  <div style="font-weight:700;color:#111827;font-size:.9rem">{s['p']} &nbsp; {s['t']}</div>
                  <div style="color:#4b5563;font-size:.83rem;margin-top:5px;line-height:1.5">→ {s['a']}</div>
                  <span style="display:inline-block;font-size:.72rem;font-weight:600;padding:2px 9px;
                    border-radius:20px;margin-top:6px;background:{chip}">Impact : {s['i']}</span>
                </div>""", unsafe_allow_html=True)

    # ── ARBORESCENCE ─────────────────────────────────────────
    def build_tree_html(df_tree, domain):
        """Construit une arborescence HTML scrollable des URLs"""
        # Liste indentée avec couleurs selon score
        lines = []
        df_sorted = df_tree.sort_values("url")
        prev_parts = []
        for _, row in df_sorted.iterrows():
            url = row["url"]
            path = url.replace(domain,"").strip("/") or "/"
            parts = path.split("/")
            depth = len(parts) - 1
            label = parts[-1] or "/"
            sc = row.get("seo_score", 0)
            st_code = row.get("status", 200)
            sc_color = "#059669" if sc >= 70 else ("#f8ba07" if sc >= 40 else "#ef4444")
            st_color = "#059669" if st_code == 200 else "#ef4444"
            indent = "&nbsp;" * (depth * 4)
            # Icône selon type
            ftype = row.get("file_type","html")
            icon = {"pdf":"📄","image":"🖼","doc":"📋","media":"🎬"}.get(ftype,"📃")
            lines.append(
                f'<div style="padding:2px 0;font-size:.82rem;font-family:monospace;white-space:nowrap">'
                f'{indent}'
                f'<span style="color:#6b7280">{"└─ " if depth>0 else ""}</span>'
                f'{icon} '
                f'<a href="{url}" target="_blank" style="color:#111827;text-decoration:none">{label or "/"}</a>'
                f' &nbsp;<span style="background:{st_color}22;color:{st_color};padding:1px 6px;border-radius:10px;font-size:.75rem">{st_code}</span>'
                f' <span style="background:{sc_color}22;color:{sc_color};padding:1px 6px;border-radius:10px;font-size:.75rem">{sc}/100</span>'
                f'</div>'
            )
        return "\n".join(lines)

    # ── ONGLETS
    if expert_mode:
        tabs=st.tabs(["🔴 Priorités","🌳 Arbo","📝 SEO On-page","🔎 SERP Preview","📄 Fichiers","🔗 Liens","⏱ Vitesse","🔁 Duplicats","⚠️ Cannibalisation","🌐 OG & Schema","🗺 Maillage","🤖 Robots & Sitemap","🔑 Mots-clés","♿ Accessibilité","⚡ Core Web Vitals","🔒 Sécurité","📊 Historique"])
        t_prio,t_tree,t_seo,t_serp,t_files,t_out,t_speed,t_dup,t_cannibal,t_og,t_mail,t_robots,t_kw,t_access,t_cwv,t_sec,t_hist=tabs
    else:
        tabs=st.tabs(["🔴 Priorités","🌳 Arbo","📄 Fichiers","🔗 Liens","🔁 Duplicats","🤖 Robots & Sitemap","📊 Historique"])
        t_prio,t_tree,t_files,t_out,t_dup,t_robots,t_hist=tabs
        t_seo=None; t_serp=None; t_speed=None; t_og=None; t_mail=None; t_kw=None; t_cwv=None; t_cannibal=None; t_access=None; t_sec=None

    with t_prio:
        st.markdown("#### Filtres")
        fc1,fc2,fc3,fc4 = st.columns(4)
        with fc1:
            filter_status = st.multiselect("Statut HTTP",
                options=sorted(df["status"].unique().tolist()),
                default=[], key="f_status",
                placeholder="Tous les statuts")
        with fc2:
            filter_score_max = st.slider("Score SEO max", 0, 100, 100, key="f_score")
        with fc3:
            filter_search = st.text_input("🔍 Rechercher dans les URLs", "", key="f_url")
        with fc4:
            filter_issue = st.text_input("🔍 Rechercher un problème", "", key="f_issue")

        df_filt = df.copy()
        if filter_status:
            df_filt = df_filt[df_filt["status"].isin(filter_status)]
        if "seo_score" in df_filt.columns:
            df_filt = df_filt[df_filt["seo_score"] <= filter_score_max]
        if filter_search:
            df_filt = df_filt[df_filt["url"].str.contains(filter_search, case=False, na=False)]
        if filter_issue and "issues" in df_filt.columns:
            df_filt = df_filt[df_filt["issues"].str.contains(filter_issue, case=False, na=False)]

        st.caption(f"{len(df_filt)} résultat(s) sur {total}")
        cols_c = [c for c in ["url","status","file_type","seo_score","load_time_ms","word_count","issues"] if c in df_filt.columns]
        st.dataframe(
            df_filt[cols_c].sort_values("seo_score") if "seo_score" in df_filt.columns else df_filt[cols_c],
            use_container_width=True, height=420
        )

        st.markdown("#### Top problèmes fréquents")
        all_iss=[]
        for iss in df.get("issues", pd.Series()).dropna():
            for i in str(iss).split(" - "):
                if i.strip(): all_iss.append(i.strip())
        if all_iss:
            st.dataframe(pd.DataFrame(Counter(all_iss).most_common(10),
                columns=["Problème","Occurrences"]), use_container_width=True)

    with t_tree:
        st.markdown("#### Arborescence du site crawlé")
        tc1, tc2, tc3 = st.columns(3)
        with tc1:
            tree_filter_score = st.slider("Afficher pages avec score ≤", 0, 100, 100, key="tree_score")
        with tc2:
            tree_filter_status = st.multiselect("Statuts", sorted(df["status"].unique().tolist()), default=[], key="tree_status", placeholder="Tous")
        with tc3:
            tree_search = st.text_input("🔍 Filtrer les URLs", "", key="tree_search")

        df_tree = df.copy()
        if "seo_score" in df_tree.columns:
            df_tree = df_tree[df_tree["seo_score"] <= tree_filter_score]
        if tree_filter_status:
            df_tree = df_tree[df_tree["status"].isin(tree_filter_status)]
        if tree_search:
            df_tree = df_tree[df_tree["url"].str.contains(tree_search, case=False, na=False)]

        st.caption(f"{len(df_tree)} URLs affichées")

        if len(df_tree) > 0:
            tree_html = build_tree_html(df_tree, domain)
            st.markdown(
                f'<div style="background:white;border-radius:10px;padding:1rem 1.2rem;'
                f'max-height:600px;overflow-y:auto;overflow-x:auto;'
                f'box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #e5e7eb">'
                f'{tree_html}'
                f'</div>',
                unsafe_allow_html=True
            )
            st.caption("🟢 Score ≥ 70  🟡 Score 40-69  🔴 Score < 40  -  Cliquez sur une URL pour l'ouvrir")
        else:
            st.info("Aucune URL ne correspond aux filtres.")

    with t_files:
        st.markdown("#### Fichiers détectés sur le site")
        if df_files.empty:
            st.info("Aucun fichier non-HTML trouvé - ou les types n'étaient pas cochés dans la sidebar.")
        else:
            type_counts = df_files["file_type"].value_counts().reset_index()
            type_counts.columns = ["Type","Nombre"]
            st.dataframe(type_counts, use_container_width=True)
            st.markdown("---")
            # Filtres fichiers
            ff1, ff2 = st.columns(2)
            with ff1:
                types_dispo = df_files["file_type"].unique().tolist()
                selected_type = st.selectbox("Type de fichier", ["Tous"] + types_dispo)
            with ff2:
                file_search = st.text_input("🔍 Rechercher", "", key="file_search")
            df_show = df_files.copy()
            if selected_type != "Tous":
                df_show = df_show[df_show["file_type"] == selected_type]
            if file_search:
                df_show = df_show[df_show["url"].str.contains(file_search, case=False, na=False)]
            st.caption(f"{len(df_show)} fichier(s)")
            st.dataframe(df_show.rename(columns={
                "url":"URL","file_type":"Type","file_size_kb":"Taille (Ko)","status":"Statut"}),
                use_container_width=True, height=450)

    with t_out:
        st.markdown("#### Liens sortants (off-page) - domaines externes")
        if df_outbound.empty:
            st.info("Aucun lien sortant détecté.")
        else:
            top_dom = df_outbound["domaine"].value_counts().reset_index()
            top_dom.columns = ["Domaine externe","Nb de liens"]
            c1, c2 = st.columns([1, 2])
            with c1:
                st.markdown(f"**{len(top_dom)} domaines uniques**")
                st.dataframe(top_dom.head(30), use_container_width=True, height=450)
            with c2:
                st.markdown("**Tous les liens sortants**")
                od1, od2 = st.columns(2)
                with od1:
                    search_dom = st.text_input("Filtrer domaine", "", key="dom_filter")
                with od2:
                    search_anc = st.text_input("Filtrer ancre", "", key="anc_filter")
                df_out_show = df_outbound.copy()
                if search_dom:
                    df_out_show = df_out_show[df_out_show["domaine"].str.contains(search_dom, case=False, na=False)]
                if search_anc:
                    df_out_show = df_out_show[df_out_show["ancre"].str.contains(search_anc, case=False, na=False)]
                st.caption(f"{len(df_out_show)} lien(s)")
                st.dataframe(df_out_show.rename(columns={
                    "page_source":"Page source","url_cible":"URL cible",
                    "domaine":"Domaine","ancre":"Ancre"}),
                    use_container_width=True, height=400)

    if expert_mode:
        with t_seo:
            st.markdown("#### Analyse SEO On-Page complète")
            sf1, sf2, sf3 = st.columns(3)
            with sf1:
                score_max = st.slider("Score SEO max", 0, 100, 100, key="seo_score_max")
            with sf2:
                seo_search = st.text_input("🔍 URL contient", "", key="seo_url_search")
            with sf3:
                sort_col = st.selectbox("Trier par", ["seo_score","load_time_ms","word_count","title_len"], key="seo_sort")

            cols_seo = [c for c in ["url","status","title","title_len","h1","h1_count",
                "meta_description","meta_desc_len","canonical","word_count",
                "h2_count","h3_count","depth","seo_score"] if c in df.columns]
            df_seo = df[cols_seo].copy()
            if "seo_score" in df_seo.columns:
                df_seo = df_seo[df_seo["seo_score"] <= score_max]
            if seo_search:
                df_seo = df_seo[df_seo["url"].str.contains(seo_search, case=False, na=False)]
            if sort_col in df_seo.columns:
                df_seo = df_seo.sort_values(sort_col)
            st.caption(f"{len(df_seo)} pages")
            st.dataframe(df_seo, use_container_width=True, height=550)

        # ── SERP PREVIEW ──────────────────────────────────────────
        with t_serp:
            st.markdown("#### 🔎 Aperçu SERP - rendu Google")
            st.caption("Visualise exactement comment Google affichera le title et la méta de chaque page")

            html_pages_serp = df[df["file_type"]=="html"]["url"].tolist() if "file_type" in df.columns else df["url"].tolist()

            # Filtre rapide pages problématiques
            sp_col1, sp_col2 = st.columns([2,1])
            with sp_col1:
                serp_url_sel = st.selectbox("Choisir une page", html_pages_serp, key="serp_url_sel")
            with sp_col2:
                serp_filter = st.selectbox("Filtre rapide", ["Toutes","Title trop long (>60)","Title trop court (<30)","Méta manquante","Méta trop longue (>160)"], key="serp_filter")

            # Aperçu page sélectionnée
            row_serp = df[df["url"]==serp_url_sel].iloc[0] if not df[df["url"]==serp_url_sel].empty else None
            if row_serp is not None:
                st.markdown("**Aperçu de la page sélectionnée :**")
                st.markdown(serp_preview_html(
                    row_serp.get("title",""),
                    row_serp.get("meta_description",""),
                    row_serp["url"]
                ), unsafe_allow_html=True)

            st.markdown("---")
            st.markdown("**Vue globale - toutes les pages**")

            # Appliquer filtre
            df_serp = df[df["file_type"]=="html"].copy() if "file_type" in df.columns else df.copy()
            if serp_filter=="Title trop long (>60)": df_serp=df_serp[df_serp["title_len"]>60]
            elif serp_filter=="Title trop court (<30)": df_serp=df_serp[(df_serp["title_len"]>0)&(df_serp["title_len"]<30)]
            elif serp_filter=="Méta manquante": df_serp=df_serp[df_serp["meta_description"]==""]
            elif serp_filter=="Méta trop longue (>160)": df_serp=df_serp[df_serp["meta_desc_len"]>160]

            # Tableau enrichi
            def title_status(row):
                l=row.get("title_len",0)
                if l==0: return "❌ Manquant"
                if l>60: return f"⚠️ Trop long ({l})"
                if l<30: return f"⚠️ Trop court ({l})"
                return f"✅ OK ({l})"
            def meta_status(row):
                l=row.get("meta_desc_len",0)
                if l==0: return "❌ Manquante"
                if l>160: return f"⚠️ Trop longue ({l})"
                if l<70: return f"⚠️ Trop courte ({l})"
                return f"✅ OK ({l})"

            df_serp["Statut Title"]=df_serp.apply(title_status,axis=1)
            df_serp["Statut Méta"]=df_serp.apply(meta_status,axis=1)
            cols_show=["url","title","Statut Title","meta_description","Statut Méta"]
            cols_show=[c for c in cols_show if c in df_serp.columns]
            st.caption(f"{len(df_serp)} pages")
            st.dataframe(df_serp[cols_show], use_container_width=True, height=420)

        with t_speed:
            st.markdown("#### Temps de chargement")
            df_html = df[df["file_type"]=="html"] if "file_type" in df.columns else df
            if "load_time_ms" in df_html.columns:
                spf1, spf2 = st.columns(2)
                with spf1:
                    speed_min = st.number_input("Temps min (ms)", 0, 30000, 0, 100, key="sp_min")
                with spf2:
                    speed_search = st.text_input("🔍 URL contient", "", key="sp_search")
                df_sp = df_html[["url","load_time_ms","status"]].copy()
                df_sp = df_sp[df_sp["load_time_ms"] >= speed_min]
                if speed_search:
                    df_sp = df_sp[df_sp["url"].str.contains(speed_search, case=False, na=False)]
                df_sp["Niveau"] = df_sp["load_time_ms"].apply(
                    lambda x: "🔴 Lent (>3s)" if x>3000 else ("🟡 Moyen (1-3s)" if x>1000 else "🟢 Rapide (<1s)"))
                df_sp = df_sp.sort_values("load_time_ms", ascending=False)
                st.caption(f"{len(df_sp)} pages - Temps moyen : {int(df_html['load_time_ms'].mean())} ms")
                st.dataframe(df_sp, use_container_width=True, height=500)
                n_slow = len(df_html[df_html["load_time_ms"]>3000])
                if n_slow: st.warning(f"⚠️ {n_slow} page(s) chargent en plus de 3 secondes")

        with t_og:
            st.markdown("#### Open Graph, Twitter Card & Données structurées")
            cols_og = [c for c in ["url","og_title","og_description","og_image","has_schema","hreflang_count"] if c in df.columns]
            df_og = df[cols_og].copy() if cols_og else pd.DataFrame()
            if not df_og.empty:
                no_og = len(df_og[df_og["og_title"]==""]) if "og_title" in df_og.columns else 0
                has_schema_n = len(df_og[df_og["has_schema"]==True]) if "has_schema" in df_og.columns else 0
                c1,c2,c3 = st.columns(3)
                c1.metric("Sans og:title", no_og)
                c2.metric("Avec Schema.org", has_schema_n)
                c3.metric("Pages hreflang", len(df_og[df_og["hreflang_count"]>0]) if "hreflang_count" in df_og.columns else 0)
                og_filter = st.selectbox("Filtrer", ["Toutes","Sans og:title","Avec Schema.org","Sans Schema.org"], key="og_filter")
                df_og_show = df_og.copy()
                if og_filter == "Sans og:title" and "og_title" in df_og_show.columns:
                    df_og_show = df_og_show[df_og_show["og_title"]==""]
                elif og_filter == "Avec Schema.org" and "has_schema" in df_og_show.columns:
                    df_og_show = df_og_show[df_og_show["has_schema"]==True]
                elif og_filter == "Sans Schema.org" and "has_schema" in df_og_show.columns:
                    df_og_show = df_og_show[df_og_show["has_schema"]==False]
                st.dataframe(df_og_show, use_container_width=True, height=450)

        with t_mail:
            st.markdown("#### Maillage interne")
            link_counts = {}
            for _, row in df.iterrows():
                for lnk in row.get("internal_links_list",[]):
                    link_counts[lnk] = link_counts.get(lnk,0)+1
            if link_counts:
                df_lk = pd.DataFrame(sorted(link_counts.items(), key=lambda x:-x[1]),
                    columns=["URL cible","Liens entrants"])
                ml_search = st.text_input("🔍 Filtrer", "", key="ml_search")
                if ml_search:
                    df_lk = df_lk[df_lk["URL cible"].str.contains(ml_search, case=False, na=False)]
                st.dataframe(df_lk, use_container_width=True, height=400)
            orphans = set(df["url"].tolist()) - set(link_counts.keys())
            if orphans:
                st.markdown(f"#### Pages orphelines ({len(orphans)})")
                st.dataframe(pd.DataFrame(list(orphans), columns=["URL orpheline"]), use_container_width=True)
            else:
                st.success("✅ Aucune page orpheline !")

    with t_dup:
        st.markdown("#### Contenus dupliqués")
        if not dups:
            st.success("✅ Aucun contenu dupliqué !")
        else:
            st.warning(f"⚠️ {len(dups)} doublon(s) détecté(s)")
            df_dups = pd.DataFrame(dups)
            dup_search = st.text_input("🔍 Filtrer", "", key="dup_search")
            if dup_search:
                df_dups = df_dups[df_dups["url_1"].str.contains(dup_search,case=False,na=False) | df_dups["url_2"].str.contains(dup_search,case=False,na=False)]
            st.dataframe(df_dups.rename(columns={"url_1":"Page 1","url_2":"Page 2","type":"Type"}), use_container_width=True)

    if expert_mode and t_cannibal:
        with t_cannibal:
            st.markdown("#### ⚠️ Cannibalisation de mots-clés")
            st.caption("Détecte les pages qui ciblent les mêmes mots-clés principaux - risque de concurrence interne")
            with st.spinner("Analyse de la cannibalisation…"):
                cannibal_data=detect_cannibalization(df)
            if not cannibal_data:
                st.success("✅ Aucune cannibalisation détectée parmi les pages analysées.")
            else:
                st.warning(f"⚠️ {len(cannibal_data)} mot(s)-clé(s) en cannibalisation potentielle")
                df_cannibal=pd.DataFrame(cannibal_data)
                cannibal_min=st.slider("Nb minimum de pages en conflit",2,10,2,key="cannibal_min")
                df_cannibal=df_cannibal[df_cannibal["nb_pages"]>=cannibal_min]
                st.dataframe(df_cannibal.rename(columns={"mot_clé":"Mot-clé","nb_pages":"Pages en conflit","pages":"URLs concernées"}),
                    use_container_width=True, height=400)
                with st.expander("ℹ️ Comment corriger la cannibalisation ?"):
                    st.markdown("""
**Cannibalisation = plusieurs pages ciblent le même mot-clé** → Google ne sait pas laquelle classer.

**Solutions :**
- **Fusionner** les pages les moins performantes dans la page principale
- **Différencier** le ciblage : chaque page cible un angle unique du mot-clé
- **Rediriger** en 301 les pages doublons vers la page canonique
- **Utiliser canonical** si les pages doivent coexister pour des raisons techniques
                    """)

    if expert_mode:
        # ── ROBOTS & SITEMAP ─────────────────────────────────────
        with t_robots:
            st.markdown("#### 🤖 Robots.txt")
            if robots_data:
                rb1, rb2, rb3, rb4 = st.columns(4)
                rb1.metric("Trouvé", "✅ Oui" if robots_data.get("found") else "❌ Non")
                rb2.metric("Règles Disallow", robots_data.get("disallow_count", 0))
                rb3.metric("Règles Allow", robots_data.get("allow_count", 0))
                rb4.metric("Sitemaps référencés", len(robots_data.get("sitemap_refs", [])))

                if robots_data.get("issues"):
                    for issue in robots_data["issues"]:
                        st.warning(f"⚠️ {issue}")

                if robots_data.get("disallowed_paths"):
                    st.markdown("**Chemins bloqués :**")
                    st.dataframe(pd.DataFrame(robots_data["disallowed_paths"], columns=["Chemin Disallow"]),
                        use_container_width=True, height=200)

                if robots_data.get("user_agents"):
                    st.markdown(f"**User-agents détectés :** {', '.join(robots_data['user_agents'])}")

                with st.expander("📄 Contenu brut robots.txt"):
                    st.code(robots_data.get("raw","(vide)"), language="text")
            else:
                st.info("Relance un audit pour analyser robots.txt")

            st.markdown("---")
            st.markdown("#### 🗺️ Sitemap.xml")
            if sitemap_data:
                sm1, sm2, sm3 = st.columns(3)
                sm1.metric("Trouvé", "✅ Oui" if sitemap_data.get("found") else "❌ Non")
                sm2.metric("URLs dans le sitemap", sitemap_data.get("url_count", 0))
                sm3.metric("Type", "Sitemap Index" if sitemap_data.get("is_index") else "Sitemap simple")

                if sitemap_data.get("issues"):
                    for issue in sitemap_data["issues"]:
                        st.warning(f"⚠️ {issue}")

                if sitemap_data.get("found"):
                    # Comparaison sitemap vs crawl
                    crawled_urls = set(df["url"].tolist())
                    sitemap_urls = set(sitemap_data.get("urls_sample", []))
                    not_crawled = sitemap_urls - crawled_urls
                    if not_crawled:
                        st.markdown(f"**URLs dans sitemap non crawlées ({len(not_crawled)}) :**")
                        st.dataframe(pd.DataFrame(list(not_crawled), columns=["URL"]), use_container_width=True)

                if sitemap_data.get("sub_sitemaps"):
                    st.markdown("**Sous-sitemaps détectés :**")
                    st.dataframe(pd.DataFrame(sitemap_data["sub_sitemaps"], columns=["URL sous-sitemap"]),
                        use_container_width=True)

                if sitemap_data.get("urls_sample"):
                    st.markdown("**Extrait des URLs :**")
                    st.dataframe(pd.DataFrame(sitemap_data["urls_sample"], columns=["URL"]),
                        use_container_width=True)
            else:
                st.info("Relance un audit pour analyser sitemap.xml")

        # ── MOTS-CLÉS ─────────────────────────────────────────────
        with t_kw:
            st.markdown("#### 🔑 Analyse des mots-clés par page")
            st.caption("Densité des mots-clés principaux + présence dans Title, H1, Méta-description")

            # Sélecteur de page
            html_pages = df[df["file_type"]=="html"]["url"].tolist() if "file_type" in df.columns else df["url"].tolist()
            if html_pages:
                kw_url = st.selectbox("Choisir une page à analyser", html_pages, key="kw_url_select")
                row_kw = df[df["url"]==kw_url].iloc[0] if not df[df["url"]==kw_url].empty else None
                if row_kw is not None:
                    kw_data = analyze_page_keywords(
                        row_kw.to_dict(),
                        row_kw.get("title",""),
                        row_kw.get("h1",""),
                        row_kw.get("meta_description","")
                    )
                    if kw_data:
                        df_kw = pd.DataFrame(kw_data)
                        st.dataframe(df_kw, use_container_width=True, height=350)

                        # Analyse globale
                        st.markdown("---")
                        st.markdown("#### 📊 Top 10 mots-clés sur l'ensemble du site")
                        all_text = " ".join(df["body_text_sample"].fillna("").tolist())
                        site_kws = extract_keywords(all_text, top_n=10)
                        if site_kws:
                            df_site_kw = pd.DataFrame(site_kws, columns=["Mot-clé","Occurrences (échantillon)"])
                            st.dataframe(df_site_kw, use_container_width=True)
                    else:
                        st.info("Pas assez de contenu textuel pour cette page.")
            else:
                st.info("Aucune page HTML disponible.")

        # ── CORE WEB VITALS ───────────────────────────────────────
    with t_hist:
        st.markdown("#### Historique des audits - comparaison avant/après")
        history = st.session_state["history"]
        if len(history) < 2:
            st.info("Lance l'audit 2 fois sur le même site pour voir l'évolution.")
            if history:
                st.dataframe(pd.DataFrame(history), use_container_width=True)
        else:
            st.dataframe(pd.DataFrame(history), use_container_width=True)
            last=history[-1]; prev=history[-2]
            st.markdown("#### Delta entre les 2 derniers audits")
            e1,e2,e3,e4 = st.columns(4)
            e1.metric("Score", last["score"], delta=last["score"]-prev["score"])
            e2.metric("Erreurs", last["errors"], delta=last["errors"]-prev["errors"], delta_color="inverse")
            e3.metric("Sans title", last["no_title"], delta=last["no_title"]-prev["no_title"], delta_color="inverse")
            e4.metric("Sans méta", last["no_meta"], delta=last["no_meta"]-prev["no_meta"], delta_color="inverse")

        # Historique CWV dans le même onglet
        if cwv_history:
            st.markdown("---")
            st.markdown("#### ⚡ Évolution des scores PageSpeed dans le temps")
            df_cwv_hist = pd.DataFrame(cwv_history)
            cols_cwv = [c for c in ["date","domain","mobile_score","desktop_score","mobile_lcp","desktop_lcp","accessibility_score","render_blocking"] if c in df_cwv_hist.columns]
            st.dataframe(df_cwv_hist[cols_cwv].rename(columns={
                "date":"Date","domain":"Domaine","mobile_score":"Perf. Mobile",
                "desktop_score":"Perf. Desktop","mobile_lcp":"LCP Mobile",
                "desktop_lcp":"LCP Desktop","accessibility_score":"Accessibilité",
                "render_blocking":"Ressources bloquantes"
            }), use_container_width=True)
            if len(cwv_history)>=2:
                last_cwv=cwv_history[-1]; prev_cwv=cwv_history[-2]
                st.markdown("**Delta CWV - dernier audit vs précédent**")
                cw1,cw2,cw3 = st.columns(3)
                if "mobile_score" in last_cwv and "mobile_score" in prev_cwv:
                    cw1.metric("Perf. Mobile", last_cwv["mobile_score"],
                        delta=last_cwv["mobile_score"]-prev_cwv["mobile_score"])
                if "desktop_score" in last_cwv and "desktop_score" in prev_cwv:
                    cw2.metric("Perf. Desktop", last_cwv["desktop_score"],
                        delta=last_cwv["desktop_score"]-prev_cwv["desktop_score"])
                if "accessibility_score" in last_cwv and last_cwv["accessibility_score"] and prev_cwv.get("accessibility_score"):
                    cw3.metric("Accessibilité", last_cwv["accessibility_score"],
                        delta=last_cwv["accessibility_score"]-prev_cwv["accessibility_score"])

    # ── ROBOTS & SITEMAP (mode novice - toujours visible)
    with t_robots:
        st.markdown("#### 🤖 Analyse robots.txt")
        if robots_data:
            found_rb = robots_data.get("found", False)
            rb_color = "#059669" if found_rb else "#ef4444"
            rb_found_val = "✅ Trouvé" if found_rb else "❌ Absent"
            st.markdown(f"""
            <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:1rem">
              <div style="background:white;border-radius:12px;padding:.9rem 1rem;box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB;border-left:3px solid {rb_color}">
                <div style="font-size:1.4rem;font-weight:800;color:#111827">{rb_found_val}</div>
                <div style="font-size:.75rem;color:#6B7280;margin-top:3px;text-transform:uppercase;letter-spacing:.04em">Statut robots.txt</div>
              </div>
              <div style="background:white;border-radius:12px;padding:.9rem 1rem;box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB;border-left:3px solid #2563EB">
                <div style="font-size:1.4rem;font-weight:800;color:#111827">{robots_data.get("disallow_count", 0)}</div>
                <div style="font-size:.75rem;color:#6B7280;margin-top:3px;text-transform:uppercase;letter-spacing:.04em">Règles Disallow</div>
              </div>
              <div style="background:white;border-radius:12px;padding:.9rem 1rem;box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB;border-left:3px solid #6B7280">
                <div style="font-size:1.4rem;font-weight:800;color:#111827">{len(robots_data.get("user_agents", []))}</div>
                <div style="font-size:.75rem;color:#6B7280;margin-top:3px;text-transform:uppercase;letter-spacing:.04em">User-agents</div>
              </div>
              <div style="background:white;border-radius:12px;padding:.9rem 1rem;box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB;border-left:3px solid #2563EB">
                <div style="font-size:1.4rem;font-weight:800;color:#111827">{len(robots_data.get("sitemap_refs", []))}</div>
                <div style="font-size:.75rem;color:#6B7280;margin-top:3px;text-transform:uppercase;letter-spacing:.04em">Refs sitemap</div>
              </div>
            </div>
            """, unsafe_allow_html=True)
            if robots_data.get("issues"):
                for issue in robots_data["issues"]:
                    st.markdown(f"""<div style="background:#fffbeb;border:1px solid #f59e0b;border-radius:8px;padding:.7rem 1rem;margin-bottom:.5rem;color:#92400e;font-size:.85rem">⚠️ {issue}</div>""", unsafe_allow_html=True)
            if robots_data.get("disallowed_paths"):
                st.markdown("**Chemins bloqués :**")
                st.dataframe(pd.DataFrame(robots_data["disallowed_paths"],columns=["Chemin Disallow"]),
                    use_container_width=True, height=200)
            if robots_data.get("sitemap_refs"):
                st.markdown("**Sitemaps référencés :**")
                for s in robots_data["sitemap_refs"]: st.code(s)
            if robots_data.get("raw"):
                with st.expander("📄 Contenu brut robots.txt"):
                    st.code(robots_data["raw"], language="text")
        else:
            st.markdown("""<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:.8rem 1rem;color:#1e40af;font-size:.85rem">ℹ️ Relancez un audit pour analyser robots.txt</div>""", unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("#### 🗺 Analyse sitemap.xml")
        if sitemap_data:
            found_sm = sitemap_data.get("found", False)
            sm_color = "#059669" if found_sm else "#ef4444"
            sm_found_val = "✅ Trouvé" if found_sm else "❌ Absent"
            sm_type = "Index" if sitemap_data.get("is_index") else "Standard"
            st.markdown(f"""
            <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:1rem">
              <div style="background:white;border-radius:12px;padding:.9rem 1rem;box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB;border-left:3px solid {sm_color}">
                <div style="font-size:1.4rem;font-weight:800;color:#111827">{sm_found_val}</div>
                <div style="font-size:.75rem;color:#6B7280;margin-top:3px;text-transform:uppercase;letter-spacing:.04em">Statut sitemap</div>
              </div>
              <div style="background:white;border-radius:12px;padding:.9rem 1rem;box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB;border-left:3px solid #2563EB">
                <div style="font-size:1.4rem;font-weight:800;color:#111827">{sitemap_data.get("url_count", 0)}</div>
                <div style="font-size:.75rem;color:#6B7280;margin-top:3px;text-transform:uppercase;letter-spacing:.04em">URLs indexées</div>
              </div>
              <div style="background:white;border-radius:12px;padding:.9rem 1rem;box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB;border-left:3px solid #6B7280">
                <div style="font-size:1.4rem;font-weight:800;color:#111827">{sm_type}</div>
                <div style="font-size:.75rem;color:#6B7280;margin-top:3px;text-transform:uppercase;letter-spacing:.04em">Type</div>
              </div>
            </div>
            """, unsafe_allow_html=True)
            if sitemap_data.get("issues"):
                for issue in sitemap_data["issues"]:
                    st.markdown(f"""<div style="background:#fffbeb;border:1px solid #f59e0b;border-radius:8px;padding:.7rem 1rem;margin-bottom:.5rem;color:#92400e;font-size:.85rem">⚠️ {issue}</div>""", unsafe_allow_html=True)
            if sitemap_data.get("url",""):
                st.caption(f"URL analysée : {sitemap_data['url']}")
            if sitemap_data.get("sub_sitemaps"):
                st.markdown("**Sous-sitemaps :**")
                for s in sitemap_data["sub_sitemaps"]: st.code(s)
            elif sitemap_data.get("urls_sample"):
                st.markdown("**Aperçu des URLs :**")
                for u in sitemap_data["urls_sample"]: st.code(u)
        else:
            st.markdown("""<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:.8rem 1rem;color:#1e40af;font-size:.85rem">ℹ️ Relancez un audit pour analyser le sitemap</div>""", unsafe_allow_html=True)

    # ── ONGLETS EXPERT UNIQUEMENT ─────────────────────────────
    if expert_mode:

        # ── MOTS-CLÉS
        with t_kw:
            st.markdown("#### 🔑 Analyse des mots-clés par page")
            df_html_kw = df[df["file_type"]=="html"].copy() if "file_type" in df.columns else df.copy()
            if df_html_kw.empty:
                st.info("Aucune page HTML analysée.")
            else:
                kw_url = st.selectbox("Choisir une page à analyser",
                    df_html_kw["url"].tolist(), key="kw_url_select2")
                row_kw = df_html_kw[df_html_kw["url"]==kw_url].iloc[0]
                kw_data = analyze_page_keywords(
                    row_kw.to_dict(),
                    row_kw.get("title",""),
                    row_kw.get("h1",""),
                    row_kw.get("meta_description",""))
                if kw_data:
                    st.caption(f"Top mots-clés - {int(row_kw.get('word_count',0))} mots - Flesch : {row_kw.get('flesch_score','N/A')}")
                    df_kw = pd.DataFrame(kw_data)
                    st.dataframe(df_kw, use_container_width=True, height=350)
                    no_kw_title = len([k for k in kw_data if k["dans_title"]=="❌"])
                    if no_kw_title >= 5:
                        st.warning(f"⚠️ {no_kw_title} mots-clés fréquents absents du title - pensez à optimiser.")
                else:
                    st.info("Pas assez de contenu textuel pour analyser les mots-clés.")

                st.markdown("---")
                st.markdown("#### 📖 Lisibilité du contenu (score Flesch)")
                if TEXTSTAT_OK and "flesch_score" in df.columns:
                    df_fl = df[df["flesch_score"].notna()][["url","flesch_score","word_count"]].copy()
                    df_fl["Niveau"] = df_fl["flesch_score"].apply(lambda s: flesch_label(s)[0])
                    df_fl = df_fl.sort_values("flesch_score")
                    fl1,fl2,fl3 = st.columns(3)
                    fl1.metric("Score moyen", f"{df_fl['flesch_score'].mean():.1f}/100")
                    fl2.metric("Pages difficiles (<30)", len(df_fl[df_fl["flesch_score"]<30]))
                    fl3.metric("Pages faciles (≥70)", len(df_fl[df_fl["flesch_score"]>=70]))
                    st.dataframe(df_fl.rename(columns={"url":"URL","flesch_score":"Score Flesch","word_count":"Nb mots","Niveau":"Niveau de lisibilité"}),
                        use_container_width=True, height=350)
                    with st.expander("ℹ️ Comprendre le score Flesch"):
                        st.markdown("""
| Score | Niveau | Public cible |
|---|---|---|
| 70–100 | Facile | Grand public, adolescents |
| 50–70 | Standard | Adultes - niveau correct pour le web |
| 30–50 | Difficile | Étudiants, professionnels |
| 0–30 | Très difficile | Experts, contenu académique |

**Conseil :** viser ≥ 50 pour du contenu marketing ou e-commerce.
                        """)
                else:
                    st.info("Installez `textstat` pour activer l'analyse Flesch : `pip install textstat`")

                st.markdown("---")
                st.markdown("#### Vue globale - mots-clés les plus fréquents sur tout le site")
                all_text=" ".join(df_html_kw["body_text_sample"].fillna("").tolist())
                if all_text.strip():
                    global_kw=extract_keywords(all_text, top_n=20)
                    df_gkw=pd.DataFrame(global_kw, columns=["Mot-clé","Occurrences totales"])
                    st.dataframe(df_gkw, use_container_width=True, height=350)

        # ── ACCESSIBILITÉ
        with t_access:
            st.markdown("#### ♿ Accessibilité - Lighthouse + analyse locale")
            if not ps_result:
                st.warning("⚠️ Clé API PageSpeed requise pour le score Lighthouse accessibilité.")
                st.markdown("""<div style='background:rgba(37,99,235,.08);border-radius:8px;padding:1rem;border-left:3px solid #2563EB;color:#1E40AF;font-weight:500'>
                Ajoutez votre clé API PageSpeed dans la sidebar (gratuit, 25 000 req/jour)
                </div>""", unsafe_allow_html=True)
            else:
                acc_score = ps_result.get("accessibility_score")
                if acc_score is not None:
                    acc_col = "#059669" if acc_score>=90 else ("#f8ba07" if acc_score>=50 else "#ef4444")
                    acc_lbl = "Excellent" if acc_score>=90 else ("À améliorer" if acc_score>=50 else "Problématique")
                    st.markdown(f"""<div style='background:white;border-radius:12px;padding:1.2rem 1.5rem;
                        box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB;border-left:3px solid {acc_col};
                        display:inline-block;min-width:220px;margin-bottom:1rem'>
                        <div style='font-size:2.8rem;font-weight:800;color:{acc_col}'>{acc_score}<span style='font-size:1rem;color:#6B7280'>/100</span></div>
                        <div style='font-weight:600;color:#111827'>{acc_lbl}</div>
                        <div style='font-size:.75rem;color:#6B7280;text-transform:uppercase;letter-spacing:.04em'>Score Lighthouse Accessibilité</div>
                    </div>""", unsafe_allow_html=True)

                    acc_issues = ps_result.get("accessibility_issues",[])
                    if acc_issues:
                        st.markdown("**Problèmes d'accessibilité détectés :**")
                        for issue in acc_issues:
                            st.markdown(f"""<div style='background:#fff5f5;border-left:3px solid #ef4444;
                                border-radius:6px;padding:.6rem .9rem;margin-bottom:.5rem;font-size:.83rem'>
                                <b style='color:#991b1b'>{issue.get('titre','')}</b><br>
                                <span style='color:#6b7280'>{issue.get('description','')}</span>
                            </div>""", unsafe_allow_html=True)
                    else:
                        st.success("✅ Aucun problème d'accessibilité critique détecté par Lighthouse.")
                else:
                    st.info("Score d'accessibilité non disponible dans cette réponse API.")

            st.markdown("---")
            st.markdown("#### Analyse locale - CTA vides + images sans ALT")
            acc1,acc2,acc3 = st.columns(3)
            total_cta = int(df["cta_empty"].sum()) if "cta_empty" in df.columns else 0
            total_noalt = int(df["images_no_alt"].sum()) if "images_no_alt" in df.columns else 0
            total_nonwebp = int(df["images_non_webp"].sum()) if "images_non_webp" in df.columns else 0
            acc1.metric("CTA / liens vides", total_cta, help="Boutons et liens sans texte ni aria-label")
            acc2.metric("Images sans ALT", total_noalt, help="Critère WCAG 2.1 - non-conformité A")
            acc3.metric("Images non WebP", total_nonwebp, help="Impact performance + accessibilité mobile")

            if total_cta > 0 or total_noalt > 0:
                df_acc_pages = df[((df.get("cta_empty",pd.Series(0, index=df.index))>0) |
                                   (df.get("images_no_alt",pd.Series(0, index=df.index))>0))][
                    [c for c in ["url","cta_empty","images_no_alt","images_non_webp"] if c in df.columns]
                ].sort_values("cta_empty" if "cta_empty" in df.columns else "url", ascending=False)
                st.markdown("**Pages avec problèmes d'accessibilité :**")
                st.dataframe(df_acc_pages.rename(columns={
                    "url":"URL","cta_empty":"CTA vides","images_no_alt":"Img sans ALT","images_non_webp":"Img non WebP"
                }), use_container_width=True, height=350)

            with st.expander("ℹ️ Checklist accessibilité WCAG 2.1"):
                st.markdown("""
**Niveau A (obligatoire) :**
- Toutes les images ont un attribut `alt`
- Tous les boutons et liens ont un texte ou un `aria-label`
- Les formulaires ont des labels associés
- Pas de contenu uniquement basé sur la couleur

**Niveau AA (recommandé) :**
- Contraste texte/fond ≥ 4.5:1 (normal) ou 3:1 (large)
- Navigation possible au clavier uniquement
- Pas de piège au focus

**Impact SEO :** Google utilise les signaux d'accessibilité dans son évaluation de la qualité des pages.
                """)

        # ── CORE WEB VITALS
        with t_cwv:
            st.markdown("#### ⚡ Core Web Vitals - Mobile vs Desktop")
            if not ps_result and not ps_desktop:
                st.warning("⚠️ Renseigne ta clé API PageSpeed Google dans la sidebar pour activer cette analyse.")
            else:
                cwv1,cwv2=st.columns(2)
                for col,label,data in [(cwv1,"📱 Mobile",ps_result),(cwv2,"🖥 Desktop",ps_desktop)]:
                    with col:
                        st.markdown(f"**{label}**")
                        if data and "score" in data:
                            sc=data["score"]
                            sc_col="#059669" if sc>=90 else ("#f8ba07" if sc>=50 else "#ef4444")
                            st.markdown(f"""<div style='background:white;border-radius:10px;padding:1.2rem;
                                box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB;border-left:3px solid {sc_col}'>
                                <div style='font-size:2.5rem;font-weight:800;color:{sc_col}'>{sc}<span style='font-size:1rem'>/100</span></div>
                                <div style='margin-top:.8rem;font-size:.82rem;color:#6b7280'>
                                LCP : <b>{data.get("lcp","N/A")}</b> &nbsp;|&nbsp;
                                CLS : <b>{data.get("cls","N/A")}</b> &nbsp;|&nbsp;
                                TBT : <b>{data.get("tbt","N/A")}</b>
                                </div></div>""", unsafe_allow_html=True)
                        else:
                            st.info("Pas de données")
                rb_count=ps_result.get("render_blocking_count",0)
                rb_savings=ps_result.get("render_blocking_savings_ms",0)
                rb_items=ps_result.get("render_blocking_items",[])
                if rb_count>0:
                    st.markdown("---")
                    st.markdown(f"#### 🚧 Ressources render-blocking : **{rb_count}** (économie : {rb_savings} ms)")
                    for item in rb_items: st.code(item)
                else:
                    st.success("✅ Aucune ressource render-blocking.")
                with st.expander("ℹ️ Comprendre les Core Web Vitals"):
                    st.markdown("""| Métrique | Bon | À améliorer | Mauvais |\n|---|---|---|---|\n| LCP | < 2.5s | 2.5–4s | > 4s |\n| CLS | < 0.1 | 0.1–0.25 | > 0.25 |\n| TBT | < 200ms | 200–600ms | > 600ms |""")

        # ── SÉCURITÉ & COOKIES
        with t_sec:
            st.markdown("#### 🔒 Sécurité & Cookies")
            if not security_data:
                st.info("Relancez un audit pour analyser la sécurité.")
            else:
                sec_score = security_data.get("score", 0)
                sec_col = "#059669" if sec_score>=80 else ("#f8ba07" if sec_score>=50 else "#ef4444")
                sec_lbl = "Bon" if sec_score>=80 else ("À améliorer" if sec_score>=50 else "Critique")

                sc1, sc2 = st.columns([1,3])
                with sc1:
                    st.markdown(f"""<div style='background:white;border-radius:12px;padding:1.2rem;text-align:center;
                        box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB;border-left:3px solid {sec_col}'>
                        <div style='font-size:3rem;font-weight:800;color:{sec_col}'>{sec_score}</div>
                        <div style='font-size:.8rem;color:#9ca3af'>/100</div>
                        <div style='font-weight:600;color:#374151;margin-top:4px'>{sec_lbl}</div>
                        <div style='font-size:.75rem;color:#9ca3af'>Score Sécurité</div>
                    </div>""", unsafe_allow_html=True)
                with sc2:
                    h1,h2,h3,h4 = st.columns(4)
                    h1.metric("HTTPS", "✅" if security_data.get("https") else "❌")
                    h2.metric("HSTS", "✅" if security_data.get("hsts") else "❌")
                    h3.metric("CSP", "✅" if security_data.get("csp") else "❌")
                    h4.metric("X-Frame-Options", "✅" if security_data.get("x_frame") else "❌")
                    h5,h6,h7,h8 = st.columns(4)
                    h5.metric("X-Content-Type", "✅" if security_data.get("x_content_type") else "❌")
                    h6.metric("Referrer-Policy", "✅" if security_data.get("referrer_policy") else "❌")
                    h7.metric("Permissions-Policy", "✅" if security_data.get("permissions_policy") else "❌")
                    total_cookies = len(security_data.get("cookies",[]))
                    h8.metric("Cookies détectés", total_cookies)

                # Issues
                if security_data.get("issues"):
                    st.markdown("---")
                    st.markdown("**⚠️ Problèmes détectés :**")
                    for iss in security_data["issues"]:
                        st.error(f"🔴 {iss}")

                # Recommendations
                if security_data.get("recommendations"):
                    st.markdown("---")
                    st.markdown("**💡 Recommandations :**")
                    bmap={"🔴 Critique":"#ef4444","🟠 Important":"#f97316","🟡 Normal":"#f8ba07"}
                    bgmap={"🔴 Critique":"#fff5f5","🟠 Important":"#fff7ed","🟡 Normal":"#fffbeb"}
                    for rec in security_data["recommendations"]:
                        b=bmap.get(rec["p"],"#6b7280")
                        bg=bgmap.get(rec["p"],"#f9fafb")
                        st.markdown(f"""<div style='background:{bg};border-left:4px solid {b};border-radius:8px;
                            padding:.8rem 1rem;margin-bottom:.6rem;font-size:.85rem'>
                            <b style='color:#111827'>{rec["p"]} - {rec["t"]}</b><br>
                            <span style='color:#4b5563'>→ {rec["a"]}</span><br>
                            <span style='color:#6b7280;font-size:.78rem'>Impact : {rec["impact"]}</span>
                        </div>""", unsafe_allow_html=True)

                # Détail headers
                st.markdown("---")
                with st.expander("🔍 Détail des en-têtes HTTP"):
                    headers_detail = [
                        {"En-tête": "Strict-Transport-Security (HSTS)", "Valeur": security_data.get("hsts_value","❌ Absent") or "❌ Absent"},
                        {"En-tête": "Content-Security-Policy", "Valeur": (security_data.get("csp_value","")[:80]+"…" if len(security_data.get("csp_value",""))>80 else security_data.get("csp_value","")) or "❌ Absent"},
                        {"En-tête": "X-Frame-Options", "Valeur": security_data.get("x_frame_value","❌ Absent") or "❌ Absent"},
                        {"En-tête": "X-Content-Type-Options", "Valeur": "✅ nosniff" if security_data.get("x_content_type") else "❌ Absent"},
                        {"En-tête": "Referrer-Policy", "Valeur": security_data.get("referrer_value","❌ Absent") or "❌ Absent"},
                    ]
                    st.dataframe(pd.DataFrame(headers_detail), use_container_width=True, hide_index=True)

                # Cookies
                if security_data.get("cookies"):
                    st.markdown("---")
                    st.markdown("#### 🍪 Analyse des cookies")
                    ck1,ck2,ck3 = st.columns(3)
                    total_ck = len(security_data["cookies"])
                    ck1.metric("Sans flag Secure", security_data.get("cookies_no_secure",0),
                        delta=f"-{security_data.get('cookies_no_secure',0)}" if security_data.get("cookies_no_secure",0)>0 else None,
                        delta_color="inverse")
                    ck2.metric("Sans HttpOnly", security_data.get("cookies_no_httponly",0),
                        delta=f"-{security_data.get('cookies_no_httponly',0)}" if security_data.get("cookies_no_httponly",0)>0 else None,
                        delta_color="inverse")
                    ck3.metric("Sans SameSite", security_data.get("cookies_no_samesite",0),
                        delta=f"-{security_data.get('cookies_no_samesite',0)}" if security_data.get("cookies_no_samesite",0)>0 else None,
                        delta_color="inverse")
                    df_cookies = pd.DataFrame(security_data["cookies"])
                    if not df_cookies.empty:
                        df_cookies["secure"] = df_cookies["secure"].map({True:"✅",False:"❌"})
                        df_cookies["httponly"] = df_cookies["httponly"].map({True:"✅",False:"❌"})
                        df_cookies["samesite"] = df_cookies["samesite"].apply(lambda x: f"✅ {x}" if x else "❌")
                        st.dataframe(df_cookies.rename(columns={
                            "name":"Nom","domain":"Domaine","secure":"Secure",
                            "httponly":"HttpOnly","samesite":"SameSite",
                            "expires":"Expiration","path":"Chemin"
                        }), use_container_width=True, height=300)

                # Mixed content
                if security_data.get("mixed_content_pages"):
                    st.markdown("---")
                    st.markdown(f"#### 🔀 Contenu mixte HTTP/HTTPS - {len(security_data['mixed_content_pages'])} page(s)")
                    for page in security_data["mixed_content_pages"]:
                        with st.expander(f"⚠️ {page['url']} ({page['count']} ressource(s))"):
                            for res in page["ressources"]:
                                st.code(res)

                # Insecure forms
                if security_data.get("insecure_forms"):
                    st.markdown("---")
                    st.markdown(f"#### 📝 Formulaires non sécurisés - {len(security_data['insecure_forms'])}")
                    df_forms = pd.DataFrame(security_data["insecure_forms"])
                    df_forms["has_password"] = df_forms["has_password"].map({True:"🔐 Oui",False:"Non"})
                    st.dataframe(df_forms.rename(columns={"page":"Page","action":"Action formulaire","has_password":"Champ mot de passe"}),
                        use_container_width=True)

                with st.expander("ℹ️ Guide des en-têtes de sécurité"):
                    st.markdown("""
| En-tête | Rôle | Valeur recommandée |
|---|---|---|
| **HSTS** | Force HTTPS | `max-age=31536000; includeSubDomains` |
| **CSP** | Bloque scripts non autorisés | `default-src 'self'` + exceptions |
| **X-Frame-Options** | Anti-clickjacking | `SAMEORIGIN` |
| **X-Content-Type** | Anti-MIME sniffing | `nosniff` |
| **Referrer-Policy** | Contrôle fuite URL | `strict-origin-when-cross-origin` |
| **Permissions-Policy** | Restreint APIs navigateur | `camera=(), microphone=()` |

**Cookies :**
- `Secure` : transit HTTPS uniquement
- `HttpOnly` : inaccessible via JavaScript (protection XSS)
- `SameSite=Lax/Strict` : protection CSRF
                    """)

    # EXPORTS
    st.markdown("---")
    st.markdown("### 📥 Exporter")
    ex1,ex2,ex3=st.columns(3)
    with ex1:
        excel_data=to_excel(df,df_outbound if not df_outbound.empty else None,df_files if not df_files.empty else None,security_data=security_data)
        st.download_button("📊 Excel complet (.xlsx)",excel_data,
            f"audit_{urlparse(domain).netloc}_{time.strftime('%Y%m%d')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
    with ex2:
        pdf_bytes=generate_pdf(df,domain,global_score,ps_result,dups,sugs,opr_result,security_data=security_data)
        if pdf_bytes:
            st.download_button("📄 Rapport PDF",pdf_bytes,
                f"rapport_{urlparse(domain).netloc}_{time.strftime('%Y%m%d')}.pdf","application/pdf",use_container_width=True)
    with ex3:
        export={"date":datetime.datetime.now().isoformat(),"domain":domain,"score":global_score,
                "total":total,"summary":st.session_state["history"][-1] if st.session_state["history"] else {}}
        st.download_button("💾 JSON",json.dumps(export,ensure_ascii=False,indent=2),
            f"audit_{urlparse(domain).netloc}_{time.strftime('%Y%m%d')}.json","application/json",use_container_width=True)

else:
    st.markdown("""<div style='background:white;border-radius:12px;padding:2.5rem;text-align:center;border:1px solid #E5E7EB;margin-top:1rem'>
      <svg width="64" height="64" viewBox="0 0 64 64" fill="none" xmlns="http://www.w3.org/2000/svg" style="margin-bottom:1rem">
        <circle cx="28" cy="28" r="18" stroke="#2563EB" stroke-width="3" fill="none"/>
        <line x1="41" y1="41" x2="56" y2="56" stroke="#2563EB" stroke-width="3" stroke-linecap="round"/>
        <circle cx="28" cy="28" r="10" fill="rgba(37,99,235,.08)"/>
      </svg>
      <h3 style='color:#111827;margin-bottom:.5rem;font-family:Inter,sans-serif;font-weight:700'>Entrez une URL pour commencer</h3>
      <p style='color:#6B7280;max-width:480px;margin:0 auto;font-size:.9rem;line-height:1.6'>Renseignez l'URL dans la sidebar et cliquez sur <strong>Lancer l'audit</strong>.<br>
      Toutes les fonctionnalités sont gratuites et activées.</p>
    </div>""",unsafe_allow_html=True)

    st.markdown("<div style='margin-top:1.2rem'></div>", unsafe_allow_html=True)
    fc1, fc2, fc3, fc4 = st.columns(4)
    card_style = "background:white;border-radius:12px;padding:1.3rem 1.1rem;box-shadow:0 1px 4px rgba(0,0,0,.06);border:1px solid #E5E7EB;height:160px;display:flex;flex-direction:column;justify-content:flex-start"
    with fc1:
        st.markdown(f"""<div style='{card_style}'>
        <div style='font-size:1.4rem;margin-bottom:.5rem'>📄</div>
        <div style='font-weight:700;color:#111827;font-size:.9rem;margin-bottom:.4rem'>Fichiers</div>
        <div style='color:#6B7280;font-size:.78rem;line-height:1.5'>Filtre PDF, images, docs, médias</div>
        </div>""", unsafe_allow_html=True)
    with fc2:
        st.markdown(f"""<div style='{card_style}'>
        <div style='font-size:1.4rem;margin-bottom:.5rem'>🔗</div>
        <div style='font-weight:700;color:#111827;font-size:.9rem;margin-bottom:.4rem'>Liens sortants</div>
        <div style='color:#6B7280;font-size:.78rem;line-height:1.5'>Top domaines off-page + ancres détaillées</div>
        </div>""", unsafe_allow_html=True)
    with fc3:
        st.markdown(f"""<div style='{card_style}'>
        <div style='font-size:1.4rem;margin-bottom:.5rem'>📊</div>
        <div style='font-weight:700;color:#111827;font-size:.9rem;margin-bottom:.4rem'>Open PageRank</div>
        <div style='color:#6B7280;font-size:.78rem;line-height:1.5'>Score d'autorité gratuit (100 req/jour)</div>
        </div>""", unsafe_allow_html=True)
    with fc4:
        st.markdown(f"""<div style='{card_style}'>
        <div style='font-size:1.4rem;margin-bottom:.5rem'>🌐</div>
        <div style='font-weight:700;color:#111827;font-size:.9rem;margin-bottom:.4rem'>OG & Schema</div>
        <div style='color:#6B7280;font-size:.78rem;line-height:1.5'>Open Graph + données structurées Schema.org</div>
        </div>""", unsafe_allow_html=True)
